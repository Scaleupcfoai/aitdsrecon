"""
GST Parser — Parses GSTR-1, GSTR-2B, GSTR-3B, Sales Register, Purchases Register
for Lekha AI GST Reconciliation agents.
"""

import json
import re
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


def to_serializable(obj):
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    raise TypeError(f"Type {type(obj)} not serializable")


MONTHS = ["April", "May", "June", "July", "August", "September",
          "October", "November", "December", "January", "February", "March"]
MONTH_SHORT = ["Apr", "May", "Jun", "Jul", "Aug", "Sep",
               "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]


def _parse_date(val):
    """Parse various date formats from GST returns."""
    if isinstance(val, (datetime, date)):
        return val
    if not val:
        return None
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# 1. Parse GSTR-1 (Sales return filed — .xls format)
# ---------------------------------------------------------------------------

def parse_gstr1(filepath: str) -> dict:
    """Parse GSTR-1 Summary sheet to get monthly outward supply details."""
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_name("Summary")

    result = {"source": "gstr1", "monthly": {}, "categories": {}}

    # Parse category-wise monthly data
    # Structure: Category label row, then detail rows (Invoices/Taxable/IGST/CGST/SGST/Cess)
    # Categories: (a) B2B, (b) B2C Large, (c) B2C Small, (d) Exports,
    #             (e) Nil/Exempt, (f) Credit/Debit Note Registered,
    #             (g) Credit/Debit Note Unregistered, (h)-(i) Advance, (J) HSN
    current_category = None
    category_data = {}

    for r in range(ws.nrows):
        label = str(ws.cell_value(r, 0)).strip()

        # Category headers: (a), (b), (c), etc. — also (J) for HSN
        if re.match(r"^\([a-zA-Z]\)", label):
            current_category = label
            category_data[current_category] = {"monthly": {}}
            continue

        if not current_category:
            continue

        # Parse into per-category monthly data
        cat = category_data[current_category]
        if "Total Taxable Value" in label:
            for c in range(1, 13):
                month = MONTH_SHORT[c - 1]
                val = ws.cell_value(r, c) or 0
                if month not in cat["monthly"]:
                    cat["monthly"][month] = {"taxable_value": 0, "igst": 0, "cgst": 0, "sgst": 0}
                cat["monthly"][month]["taxable_value"] = val
        elif "IGST Amount" in label:
            for c in range(1, 13):
                month = MONTH_SHORT[c - 1]
                if month in cat["monthly"]:
                    cat["monthly"][month]["igst"] = ws.cell_value(r, c) or 0
        elif "CGST Amount" in label:
            for c in range(1, 13):
                month = MONTH_SHORT[c - 1]
                if month in cat["monthly"]:
                    cat["monthly"][month]["cgst"] = ws.cell_value(r, c) or 0
        elif "SGST Amount" in label:
            for c in range(1, 13):
                month = MONTH_SHORT[c - 1]
                if month in cat["monthly"]:
                    cat["monthly"][month]["sgst"] = ws.cell_value(r, c) or 0

    result["categories"] = category_data

    # Build net monthly totals: B2B + B2C Small + Credit/Debit Notes
    # This should match GSTR-3B outward values
    supply_categories = []
    for cat_key in category_data:
        cat_lower = cat_key.lower()
        if "b2b" in cat_lower or "b2c" in cat_lower or "credit" in cat_lower or "export" in cat_lower:
            supply_categories.append(cat_key)

    for cat_key in supply_categories:
        for month, vals in category_data[cat_key].get("monthly", {}).items():
            if month not in result["monthly"]:
                result["monthly"][month] = {"taxable_value": 0, "igst": 0, "cgst": 0, "sgst": 0, "cess": 0}
            result["monthly"][month]["taxable_value"] += vals.get("taxable_value", 0)
            result["monthly"][month]["igst"] += vals.get("igst", 0)
            result["monthly"][month]["cgst"] += vals.get("cgst", 0)
            result["monthly"][month]["sgst"] += vals.get("sgst", 0)

    # Remove empty months
    result["monthly"] = {m: v for m, v in result["monthly"].items()
                         if v["taxable_value"] != 0}

    wb.release_resources()
    return result


# ---------------------------------------------------------------------------
# 2. Parse GSTR-2B (ITC statement from portal — .xlsx format)
# ---------------------------------------------------------------------------

def parse_gstr2b(filepath: str) -> dict:
    """Parse GSTR-2B B2B sheet for vendor-level ITC details."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    result = {"source": "gstr2b", "invoices": [], "summary": {}}

    # Parse B2B sheet — vendor invoices
    ws = wb["B2B"]
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
        gstin = row[0].value
        if not gstin or len(str(gstin)) != 15:
            continue

        invoice_date = _parse_date(row[4].value)
        result["invoices"].append({
            "gstin": str(gstin).strip(),
            "vendor_name": str(row[1].value or "").strip(),
            "invoice_no": str(row[2].value or "").strip(),
            "invoice_type": str(row[3].value or "").strip(),
            "invoice_date": invoice_date,
            "invoice_value": row[5].value or 0,
            "place_of_supply": str(row[6].value or "").strip(),
            "reverse_charge": str(row[7].value or "").strip(),
            "rate": row[8].value or 0,
            "taxable_value": row[9].value or 0,
            "igst": row[10].value or 0,
            "cgst": row[11].value or 0,
            "sgst": row[12].value or 0,
            "cess": row[13].value or 0,
            "filing_period": str(row[14].value or "").strip(),
            "itc_available": str(row[16].value or "").strip(),
            "reason": str(row[17].value or "").strip(),
        })

    # Parse ITC Available summary
    if "ITC Available" in wb.sheetnames:
        ws_itc = wb["ITC Available"]
        for row in ws_itc.iter_rows(min_row=6, max_row=ws_itc.max_row):
            label = str(row[1].value or "").strip()
            if "B2B - Invoices" in label and "Amendment" not in label and "Debit" not in label:
                result["summary"]["b2b_invoices"] = {
                    "apr": {"igst": row[3].value or 0, "cgst": row[4].value or 0, "sgst": row[5].value or 0},
                    "may": {"igst": row[7].value or 0, "cgst": row[8].value or 0, "sgst": row[9].value or 0},
                    "jun": {"igst": row[11].value or 0, "cgst": row[12].value or 0, "sgst": row[13].value or 0},
                }

    wb.close()
    return result


# ---------------------------------------------------------------------------
# 3. Parse GSTR-3B (Monthly summary return — .xls format)
# ---------------------------------------------------------------------------

def parse_gstr3b(filepath: str) -> dict:
    """Parse GSTR-3B sections 3.1 (outward), 4 (ITC), 6 (tax paid)."""
    wb = xlrd.open_workbook(filepath)
    result = {"source": "gstr3b", "outward": {}, "itc": {}, "tax_paid": {}}

    # Section 3.1 — Outward supplies
    ws31 = wb.sheet_by_name("3.1")
    for r in range(ws31.nrows):
        label = str(ws31.cell_value(r, 0)).strip()
        if "Total Taxable Value" in label and "(a)" in str(ws31.cell_value(r - 1, 0)):
            for c in range(1, 13):
                month = MONTH_SHORT[c - 1]
                val = ws31.cell_value(r, c) or 0
                if val > 0:
                    result["outward"][month] = {"taxable_value": val, "igst": 0, "cgst": 0, "sgst": 0}
        if "IGST Amount" in label and r <= 8:
            for c in range(1, 13):
                month = MONTH_SHORT[c - 1]
                val = ws31.cell_value(r, c) or 0
                if month in result["outward"]:
                    result["outward"][month]["igst"] = val
        if "CGST Amount" in label and r <= 8:
            for c in range(1, 13):
                month = MONTH_SHORT[c - 1]
                val = ws31.cell_value(r, c) or 0
                if month in result["outward"]:
                    result["outward"][month]["cgst"] = val
        if "SGST Amount" in label and r <= 8:
            for c in range(1, 13):
                month = MONTH_SHORT[c - 1]
                val = ws31.cell_value(r, c) or 0
                if month in result["outward"]:
                    result["outward"][month]["sgst"] = val

    # Section 4 — ITC claimed
    ws4 = wb.sheet_by_name("4")
    for r in range(ws4.nrows):
        label = str(ws4.cell_value(r, 0)).strip()
        if "(5) All other ITC" in label:
            # IGST, CGST, SGST are in subsequent rows
            for rr in range(r + 1, min(r + 5, ws4.nrows)):
                sublabel = str(ws4.cell_value(rr, 0)).strip()
                for c in range(1, 4):  # Apr, May, Jun
                    month = MONTH_SHORT[c - 1]
                    val = ws4.cell_value(rr, c) or 0
                    if month not in result["itc"]:
                        result["itc"][month] = {"igst": 0, "cgst": 0, "sgst": 0}
                    if "IGST" in sublabel:
                        result["itc"][month]["igst"] = val
                    elif "CGST" in sublabel:
                        result["itc"][month]["cgst"] = val
                    elif "SGST" in sublabel:
                        result["itc"][month]["sgst"] = val
            break

    # Section 6.1 — Tax paid
    if "6.1" in wb.sheet_names():
        ws6 = wb.sheet_by_name("6.1")
        for r in range(ws6.nrows):
            label = str(ws6.cell_value(r, 0)).strip()
            if "Paid through" in label or "Tax paid" in label.lower():
                for c in range(1, 4):
                    month = MONTH_SHORT[c - 1]
                    val = ws6.cell_value(r, c) or 0
                    result["tax_paid"][month] = val

    wb.release_resources()
    return result


# ---------------------------------------------------------------------------
# 4. Parse Sales Register (Tally — .xlsx format)
# ---------------------------------------------------------------------------

def parse_sales_register(filepath: str) -> dict:
    """Parse Tally Sales Register — monthly sales by account."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {"source": "sales_register", "monthly": {}}

    for sheet_name in wb.sheetnames:
        if sheet_name == "Sales Total":
            continue
        ws = wb[sheet_name]
        month_data = {"accounts": {}, "total_sales": 0}

        for row in ws.iter_rows(min_row=12, max_row=ws.max_row):
            account = str(row[0].value or "").strip()
            if not account:
                continue
            debit = row[2].value or 0
            credit = row[3].value or 0
            net = credit - debit
            month_data["accounts"][account] = {"debit": debit, "credit": credit, "net": net}
            if "Sales Account" in account or "Cash Sale" in account:
                month_data["total_sales"] += credit

        result["monthly"][sheet_name] = month_data

    wb.close()
    return result


# ---------------------------------------------------------------------------
# 5. Parse Purchases Register (Tally — .xlsx format)
# ---------------------------------------------------------------------------

def parse_purchases_register(filepath: str) -> dict:
    """Parse Tally Purchases Register — monthly purchases by account."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {"source": "purchases_register", "monthly": {}}

    for sheet_name in wb.sheetnames:
        if sheet_name == "Purchase Total":
            continue
        ws = wb[sheet_name]
        month_data = {"accounts": {}, "total_purchases": 0}

        for row in ws.iter_rows(min_row=12, max_row=ws.max_row):
            account = str(row[0].value or "").strip()
            if not account:
                continue
            debit = row[2].value or 0
            credit = row[3].value or 0
            month_data["accounts"][account] = {"debit": debit, "credit": credit}
            if "Purchase Account" in account:
                month_data["total_purchases"] += debit

        result["monthly"][sheet_name] = month_data

    wb.close()
    return result


# ---------------------------------------------------------------------------
# Main — Parse all GST sources
# ---------------------------------------------------------------------------

def parse_all_gst(data_dir: str, output_dir: str):
    """Parse all GST source files and write to output directory."""
    data_dir = Path(data_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    results = {}

    # GSTR-1
    gstr1_path = data_dir / "GSTR 1.xls"
    if gstr1_path.exists():
        print("[GST Parser] Parsing GSTR-1...")
        results["gstr1"] = parse_gstr1(str(gstr1_path))
        print(f"  → {len(results['gstr1']['monthly'])} months of data")

    # GSTR-2B
    gstr2b_path = data_dir / "GSTR- 2B.xlsx"
    if gstr2b_path.exists():
        print("[GST Parser] Parsing GSTR-2B...")
        results["gstr2b"] = parse_gstr2b(str(gstr2b_path))
        print(f"  → {len(results['gstr2b']['invoices'])} vendor invoices")

    # GSTR-3B
    gstr3b_path = data_dir / "GSTR-3B.xls"
    if gstr3b_path.exists():
        print("[GST Parser] Parsing GSTR-3B...")
        results["gstr3b"] = parse_gstr3b(str(gstr3b_path))
        print(f"  → Outward: {len(results['gstr3b']['outward'])} months, ITC: {len(results['gstr3b']['itc'])} months")

    # Sales Register
    sales_path = data_dir / "Sales Register.xlsx"
    if sales_path.exists():
        print("[GST Parser] Parsing Sales Register...")
        results["sales"] = parse_sales_register(str(sales_path))
        print(f"  → {len(results['sales']['monthly'])} months")

    # Purchases Register
    purch_path = data_dir / "Purchases Register.xlsx"
    if purch_path.exists():
        print("[GST Parser] Parsing Purchases Register...")
        results["purchases"] = parse_purchases_register(str(purch_path))
        print(f"  → {len(results['purchases']['monthly'])} months")

    # Write parsed output
    out_file = output_dir / "parsed_gst.json"
    with open(out_file, "w") as f:
        json.dump(results, f, indent=2, default=to_serializable)
    print(f"\n[GST Parser] Wrote {out_file}")

    return results


if __name__ == "__main__":
    base = Path(__file__).parent.parent.parent
    parse_all_gst(str(base / "data" / "hpc"), str(base / "tds-recon" / "data" / "parsed"))
