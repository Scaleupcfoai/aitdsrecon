"""
TDS Checker Agent — TDS Reconciliation MVP
===========================================
Validates TDS compliance on matched entries and detects missing TDS deductions.

The Matcher Agent answers: "Which Tally entry corresponds to which Form 26 entry?"
This agent answers: "Is the TDS correct — right section, right rate, right base amount?"

Check 1: Section Validation    — Is the TDS section correct for the expense type?
Check 2: Rate Validation       — Is the TDS rate correct for entity type + section?
Check 3: Base Amount Validation — Is TDS computed on the correct base (pre-GST)?
Check 4: Threshold Validation   — Does the amount exceed the applicable threshold?
Check 5: Missing TDS Detection  — Tally expenses where TDS should have been deducted
                                  but no Form 26 entry exists.

Inputs:
  - data/parsed/parsed_form26.json
  - data/parsed/parsed_tally.json
  - data/results/match_results.json

Outputs:
  - data/results/checker_results.json
"""

import json
import math
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# TDS Rules Engine — Indian Income Tax Act
# ---------------------------------------------------------------------------

# Section → expense head mapping: which expense types belong under which section
# Key = TDS section, Value = set of expense head keywords (lowercase)
SECTION_EXPENSE_MAP = {
    "194A": {
        "keywords": {"interest", "loan"},
        "description": "Interest other than interest on securities",
    },
    "194C": {
        "keywords": {
            "freight", "carriage", "transport", "logistics", "packing",
            "printing", "stationary", "shop repair", "maintenance",
            "computer repair", "contractor",
        },
        "description": "Payment to contractors/sub-contractors",
    },
    "194H": {
        "keywords": {"brokerage", "commission"},
        "description": "Commission or brokerage",
    },
    "194J(a)": {
        "keywords": {"call centre", "technical"},
        "description": "Fee for technical services (to call centre)",
    },
    "194J(b)": {
        "keywords": {
            "professional", "consultancy", "audit", "legal", "software",
            "gst annual return", "domain",
        },
        "description": "Fee for professional/technical services",
    },
    "194Q": {
        "keywords": {"purchase"},
        "description": "Payment for purchase of goods",
    },
}

# Expense heads that could be mis-classified between sections
# advertisement can be 194C (if works contract) or 194J(b) (if professional)
AMBIGUOUS_EXPENSE_HEADS = {
    "advertisement": {
        "likely_sections": ["194C", "194J(b)"],
        "note": "Advertisement may be 194C (works contract / production) "
                "or 194J(b) (professional/creative services). "
                "Verify nature of service from invoice.",
    },
    "annual maintenance": {
        "likely_sections": ["194C", "194J(b)"],
        "note": "AMC may be 194C (if labour/facility) or 194J(b) "
                "(if technical/software). Check contract terms.",
    },
    "software": {
        "likely_sections": ["194J(b)", "194C"],
        "note": "Software expenses are typically 194J(b) for royalty/license, "
                "but 194C if it's a development contract.",
    },
}

# TDS rate table: section → entity_type → rate (%)
# entity_type: "individual_huf" or "company" (derived from PAN 4th char)
TDS_RATES = {
    "192": {
        "individual_huf": 30.0,  # Highest slab assumed for directors
        "company": 30.0,
        "default": 30.0,
    },
    "194A": {
        "individual_huf": 10.0,
        "company": 10.0,
        "default": 10.0,
    },
    "194C": {
        "individual_huf": 1.0,
        "company": 2.0,
        "default": 2.0,
    },
    "194H": {
        "individual_huf": 2.0,    # 5% before Budget 2025, now 2%
        "company": 2.0,           # 5% before Budget 2025, now 2%
        "default": 2.0,
    },
    "194J(a)": {
        "individual_huf": 2.0,
        "company": 2.0,
        "default": 2.0,
    },
    "194J(b)": {
        "individual_huf": 10.0,
        "company": 10.0,
        "default": 10.0,
    },
    "194Q": {
        "individual_huf": 0.1,
        "company": 0.1,
        "default": 0.1,
    },
}

# Threshold limits: section → {single_txn, aggregate_annual}
# Below these limits, TDS need not be deducted
TDS_THRESHOLDS = {
    "192": {
        "aggregate_annual": 250000,  # Basic exemption limit
        "single_txn": None,
        "description": "Salary below ₹2,50,000 basic exemption limit",
    },
    "194A": {
        "aggregate_annual": 5000,
        "single_txn": None,
        "description": "Interest below ₹5,000 in a year exempt",
    },
    "194C": {
        "single_txn": 30000,
        "aggregate_annual": 100000,
        "description": "Single payment ≤₹30,000 or aggregate ≤₹1,00,000 exempt",
    },
    "194H": {
        "aggregate_annual": 15000,
        "single_txn": None,
        "description": "Commission below ₹15,000 in a year exempt",
    },
    "194J(a)": {
        "aggregate_annual": 30000,
        "single_txn": None,
        "description": "Professional fees below ₹30,000 in a year exempt",
    },
    "194J(b)": {
        "aggregate_annual": 30000,
        "single_txn": None,
        "description": "Professional fees below ₹30,000 in a year exempt",
    },
    "194Q": {
        "aggregate_annual": 5000000,  # ₹50 lakh
        "single_txn": None,
        "description": "Purchase of goods below ₹50 lakh in a year exempt",
    },
}

# GST rates for reverse-computing base amounts
GST_RATES = [0.18, 0.12, 0.05, 0.28]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def entity_type_from_pan(pan: str) -> str:
    """Derive entity type from PAN 4th character.
    C = Company, P = Individual, H = HUF, F = Firm, etc."""
    if not pan or len(pan) < 4:
        return "unknown"
    fourth = pan[3].upper()
    if fourth == "C":
        return "company"
    elif fourth in ("P", "H"):
        return "individual_huf"
    elif fourth == "F":
        return "firm"  # treated as individual_huf for TDS rates
    elif fourth in ("A", "T", "B", "L", "J"):
        return "individual_huf"  # AOP, Trust, BOI, Local Auth, Artificial — individual rates
    elif fourth == "G":
        return "government"
    else:
        return "unknown"


def is_valid_pan(pan: str) -> bool:
    """Check if PAN follows the valid format: 5 letters + 4 digits + 1 letter.
    Also rejects known placeholder PANs."""
    if not pan or len(pan) != 10:
        return False
    import re
    if not re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]$', pan.upper()):
        return False
    # Reject known placeholders (all same digits, sequential)
    if pan[5:9] in ("0000", "0001", "9999"):
        return False
    return True


def pan_name_initial_matches(pan: str, vendor_name: str) -> bool:
    """Check if PAN 5th character matches the first letter of the entity name.
    For individuals (4th=P): 5th char = first letter of surname.
    For others: 5th char = first letter of entity name."""
    if not pan or len(pan) < 5 or not vendor_name:
        return True  # can't verify, assume ok
    pan5 = pan[4].upper()
    # For company/firm, check against entity name first letter
    name_initial = vendor_name.strip()[0].upper() if vendor_name.strip() else ""
    return pan5 == name_initial


def expected_rate(section: str, pan: str) -> float | None:
    """Get expected TDS rate for a section + PAN combination."""
    rates = TDS_RATES.get(section)
    if not rates:
        return None
    etype = entity_type_from_pan(pan)
    if etype == "firm":
        etype = "individual_huf"  # firms use individual rate for most sections
    return rates.get(etype, rates.get("default"))


def classify_expense_head(head: str) -> list[str]:
    """Given an expense head string, return likely TDS sections."""
    head_lower = head.lower().strip()

    # Check ambiguous first
    for keyword, info in AMBIGUOUS_EXPENSE_HEADS.items():
        if keyword in head_lower:
            return info["likely_sections"]

    # Check standard mappings
    matches = []
    for section, mapping in SECTION_EXPENSE_MAP.items():
        for keyword in mapping["keywords"]:
            if keyword in head_lower:
                matches.append(section)
                break

    return matches if matches else ["unknown"]


def normalize_name(name: str) -> str:
    """Normalize vendor name for comparison."""
    if not name:
        return ""
    n = name.lower().strip()
    for suffix in ["pvt. ltd.", "pvt ltd", "private limited", "ltd.", "ltd",
                   "llp", "lp", "inc.", "inc", "co.", "company"]:
        n = n.replace(suffix, "")
    n = re.sub(r"\(\d+\)", "", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def parse_date(d) -> datetime | None:
    if isinstance(d, datetime):
        return d
    if isinstance(d, str):
        try:
            return datetime.fromisoformat(d)
        except ValueError:
            return None
    return None


# ---------------------------------------------------------------------------
# Check 1: Section Validation
# ---------------------------------------------------------------------------

def check_section(match_entry: dict) -> dict | None:
    """Validate if the Form 26 section is correct for the matched expense type."""
    f26 = match_entry["form26_entry"]
    tally_entries = match_entry["tally_entries"]
    section = f26["section"]

    # Collect all expense heads from tally entries
    expense_heads = set()
    for t in tally_entries:
        # GST exp entries have expense_heads dict
        if "expense_heads" in t and t["expense_heads"]:
            for head in t["expense_heads"]:
                expense_heads.add(head)
        # Journal entries have account_postings
        elif "account_postings" in t and t["account_postings"]:
            for head in t["account_postings"]:
                if head not in ("Gross Total", "Value"):
                    expense_heads.add(head)
        # Infer from tally_source
        src = t.get("tally_source", "")
        if src == "journal_interest":
            expense_heads.add("Interest Paid")
        elif src == "journal_freight":
            expense_heads.add("Freight Charges")

    if not expense_heads:
        return None  # can't validate without expense info

    # For each expense head, check which sections it maps to
    all_expected_sections = set()
    ambiguous_heads = []
    for head in expense_heads:
        expected = classify_expense_head(head)
        all_expected_sections.update(expected)
        if head.lower() in AMBIGUOUS_EXPENSE_HEADS:
            ambiguous_heads.append(head)

    # If the declared section is in expected sections, it's OK
    if section in all_expected_sections:
        status = "ok"
        if ambiguous_heads:
            status = "review"  # technically valid but ambiguous
    elif "unknown" in all_expected_sections and len(all_expected_sections) == 1:
        status = "unclassified"
    else:
        status = "mismatch"

    if status == "ok":
        return None  # no finding to report

    return {
        "check": "section_validation",
        "severity": "warning" if status == "review" else ("info" if status == "unclassified" else "error"),
        "vendor": f26["vendor_name"],
        "pan": f26.get("pan", ""),
        "form26_section": section,
        "expense_heads": sorted(expense_heads),
        "expected_sections": sorted(all_expected_sections - {"unknown"}),
        "status": status,
        "message": _section_message(status, section, expense_heads, all_expected_sections, ambiguous_heads),
    }


def _section_message(status, section, heads, expected, ambiguous):
    heads_str = ", ".join(sorted(heads))
    if status == "mismatch":
        exp_str = ", ".join(sorted(expected - {"unknown"}))
        return (f"Section {section} may be incorrect for expense(s): {heads_str}. "
                f"Expected section(s): {exp_str}.")
    elif status == "review":
        amb_str = ", ".join(ambiguous)
        return (f"Ambiguous expense(s): {amb_str}. Section {section} is possible but "
                f"verify against invoice/contract nature.")
    else:
        return f"Cannot classify expense(s): {heads_str}. Manual review needed."


# ---------------------------------------------------------------------------
# Check 2: Rate Validation
# ---------------------------------------------------------------------------

def check_rate(match_entry: dict) -> list[dict]:
    """Validate TDS rate and PAN compliance for a matched entry.

    Returns a list of findings (0, 1, or 2):
    - PAN validity / Section 206AA check
    - Rate mismatch check (entity-type-aware)
    """
    f26 = match_entry["form26_entry"]
    section = f26["section"]
    pan = f26.get("pan", "")
    actual_rate = f26.get("tax_rate_pct")
    amount_paid = f26.get("amount_paid", 0)
    tax_deducted = f26.get("tax_deducted", 0)
    vendor = f26.get("vendor_name", "")

    findings = []

    # ── PAN Validation + Section 206AA ──
    if not pan or not is_valid_pan(pan):
        # Section 206AA: No valid PAN → TDS at 20% or applicable rate, whichever is higher
        section_rate = expected_rate(section, "XXXPX0000X") or 0  # default individual rate
        min_206aa_rate = max(20.0, section_rate)
        if actual_rate is not None and actual_rate < min_206aa_rate:
            short_deduction = round(amount_paid * (min_206aa_rate - actual_rate) / 100, 2)
            findings.append({
                "check": "pan_validation",
                "severity": "error",
                "vendor": vendor,
                "pan": pan or "(missing)",
                "form26_section": section,
                "actual_rate_pct": actual_rate,
                "required_rate_pct": min_206aa_rate,
                "amount_paid": amount_paid,
                "short_deduction": short_deduction,
                "message": (
                    f"Invalid/missing PAN for {vendor} — Section 206AA requires TDS at "
                    f"{min_206aa_rate}% (applied {actual_rate}%). "
                    f"Potential short deduction: ₹{short_deduction:,.0f}."
                ),
            })
        elif not pan:
            findings.append({
                "check": "pan_validation",
                "severity": "warning",
                "vendor": vendor,
                "pan": "(missing)",
                "form26_section": section,
                "message": f"PAN missing for {vendor}. Section 206AA may apply (20% minimum rate).",
            })

    else:
        # PAN is valid — check entity type vs rate
        etype = entity_type_from_pan(pan)

        if etype == "government":
            # Government entities — generally exempt, skip rate check
            return findings

        if actual_rate is None:
            return findings

        exp_rate = expected_rate(section, pan)
        if exp_rate is None:
            return findings

        # Rate mismatch check
        if abs(actual_rate - exp_rate) >= 0.01:
            if amount_paid > 0:
                computed_rate = round(tax_deducted / amount_paid * 100, 2)
            else:
                computed_rate = 0

            findings.append({
                "check": "rate_validation",
                "severity": "error",
                "vendor": vendor,
                "pan": pan,
                "entity_type": etype,
                "form26_section": section,
                "actual_rate_pct": actual_rate,
                "expected_rate_pct": exp_rate,
                "computed_rate_pct": computed_rate,
                "amount_paid": amount_paid,
                "tax_deducted": tax_deducted,
                "message": (
                    f"TDS rate mismatch: {section} on {vendor} ({etype}) — "
                    f"applied {actual_rate}% but expected {exp_rate}%."
                ),
            })

    return findings


# ---------------------------------------------------------------------------
# Check 3: Base Amount Validation
# ---------------------------------------------------------------------------

def check_base_amount(match_entry: dict) -> dict | None:
    """Validate TDS is computed on the correct base amount (pre-GST, not gross).

    For GST entries, TDS should be on base_amount (net of GST).
    If Form 26 amount_paid == Tally gross (inclusive of GST), flag it.
    """
    f26 = match_entry["form26_entry"]
    tally_entries = match_entry["tally_entries"]

    f26_amount = f26.get("amount_paid", 0)
    if f26_amount == 0:
        return None

    # Sum tally base amounts and gross amounts
    total_base = 0
    total_gross = 0
    has_gst_entries = False

    for t in tally_entries:
        if t.get("tally_source") in ("gst_exp", "purchase"):
            has_gst_entries = True
            base = t.get("amount", 0)  # in matcher, 'amount' is base for gst_exp
            gross = t.get("gross_amount", 0) or t.get("gross_total", 0) or 0
            total_base += base
            total_gross += gross
        elif t.get("tally_source") in ("journal_freight", "journal_interest"):
            # Journal entries don't have GST breakup
            total_base += t.get("amount", 0)
            total_gross += t.get("amount", 0)

    if not has_gst_entries:
        return None  # No GST component to validate

    if total_gross == 0:
        return None

    # Check: is F26 amount matching gross (wrong) instead of base (right)?
    # TDS should be on base amount (pre-GST)
    base_diff = abs(f26_amount - total_base)
    gross_diff = abs(f26_amount - total_gross)

    # F26 matches base → correct
    if base_diff <= max(1, total_base * 0.005):
        return None

    # F26 matches gross → TDS on GST-inclusive amount (incorrect)
    if gross_diff < base_diff and gross_diff <= max(1, total_gross * 0.005):
        gst_amount = total_gross - total_base
        excess_tds = round(f26.get("tax_rate_pct", 0) / 100 * gst_amount, 2)
        return {
            "check": "base_amount_validation",
            "severity": "error",
            "vendor": f26["vendor_name"],
            "pan": f26.get("pan", ""),
            "form26_section": f26["section"],
            "form26_amount": f26_amount,
            "tally_base_amount": total_base,
            "tally_gross_amount": total_gross,
            "gst_component": round(total_gross - total_base, 2),
            "excess_tds": excess_tds,
            "message": (f"TDS appears computed on GST-inclusive amount (₹{f26_amount:,}) "
                        f"instead of base (₹{total_base:,.0f}). GST component: ₹{gst_amount:,.0f}. "
                        f"Potential excess TDS: ₹{excess_tds:,.0f}."),
        }

    return None


# ---------------------------------------------------------------------------
# Check 4: Threshold Validation
# ---------------------------------------------------------------------------

def check_thresholds(match_entries: list[dict], tally_data: dict = None) -> list[dict]:
    """Validate threshold limits across all matches.

    For each vendor+section in Form 26, checks if TOTAL BOOKS EXPENSE
    for that vendor IN THAT SECTION is below the annual threshold.
    Uses classify_expense_head() to route each Tally entry to correct section.
    """
    findings = []

    # Build total books expense per (vendor_normalized, section) from Tally
    # Each entry is classified by its expense head to determine the TDS section
    books_by_vendor_section = defaultdict(float)  # (vendor_norm, section) → total

    # Map journal entry_type to section
    JOURNAL_TYPE_TO_SECTION = {
        "interest_payment": "194A",
        "freight_expense": "194C",
        "packing_expense": "194C",
        "brokerage": "194H",
        "professional_fees": "194J(b)",
        "consultancy": "194J(b)",
        "audit_fees": "194J(b)",
        "rent": "194I",
        "salary": "192",
    }

    if tally_data:
        # Journal Register — classify by entry_type, use posting amounts
        for e in tally_data.get("journal_register", {}).get("entries", []):
            vendor = normalize_name(e.get("particulars", "") or e.get("loan_party", ""))
            if not vendor:
                continue
            entry_type = e.get("entry_type", "")
            section = JOURNAL_TYPE_TO_SECTION.get(entry_type)
            if section:
                # Use posting amounts (not gross_total which can be 0 for interest entries)
                postings = e.get("account_postings", {})
                amount = abs(e.get("gross_total", 0) or 0)
                if amount == 0:
                    # Sum non-meta postings
                    for k, v in postings.items():
                        if k.lower() not in ("gross total", "value", "tds payable"):
                            amount = max(amount, abs(v or 0))
                books_by_vendor_section[(vendor, section)] += amount

        # GST Exp Register — classify by expense heads
        for e in tally_data.get("purchase_gst_exp_register", {}).get("entries", []):
            vendor = normalize_name(e.get("particulars", ""))
            if not vendor:
                continue
            heads = e.get("expense_heads", {})
            for head in heads:
                sections = classify_expense_head(head)
                for sec in sections:
                    if sec != "unknown":
                        books_by_vendor_section[(vendor, sec)] += abs(heads.get(head, 0))

        # Purchase Register — all are 194Q
        for e in tally_data.get("purchase_register", {}).get("entries", []):
            vendor = normalize_name(e.get("particulars", ""))
            if vendor:
                books_by_vendor_section[(vendor, "194Q")] += abs(e.get("gross_total", 0) or 0)

    # Group Form 26 by (vendor_normalized, section)
    vendor_section_totals = defaultdict(lambda: {
        "f26_amount": 0, "entries": [], "vendor_name": "", "pan": "",
    })

    for m in match_entries:
        f26 = m["form26_entry"]
        key = (normalize_name(f26["vendor_name"]), f26["section"])
        vs = vendor_section_totals[key]
        vs["f26_amount"] += f26.get("amount_paid", 0)
        vs["entries"].append(f26)
        vs["vendor_name"] = f26["vendor_name"]
        vs["pan"] = f26.get("pan", "")

    for (vendor_norm, section), vs in vendor_section_totals.items():
        thresholds = TDS_THRESHOLDS.get(section)
        if not thresholds:
            continue

        f26_total = vs["f26_amount"]
        agg_limit = thresholds.get("aggregate_annual")

        if not agg_limit:
            continue

        # Use total books expense for this vendor IN THIS SECTION
        total_books = books_by_vendor_section.get((vendor_norm, section), f26_total)

        # If total books expense is ABOVE threshold → TDS is mandatory, no finding needed
        if total_books >= agg_limit:
            continue

        # Total books expense is below threshold → TDS is voluntary
        findings.append({
            "check": "threshold_validation",
            "severity": "info",
            "vendor": vs["vendor_name"],
            "pan": vs["pan"],
            "form26_section": section,
            "aggregate_amount": f26_total,
            "books_total": round(total_books, 2),
            "threshold_annual": agg_limit,
            "num_entries": len(vs["entries"]),
            "status": "below_threshold_but_deducted",
            "message": (f"{section} for {vs['vendor_name']}: total expenses in books ₹{total_books:,.0f} "
                        f"is below annual threshold ₹{agg_limit:,}. "
                        f"TDS deduction is voluntary (not wrong if deducted)."),
        })

    return findings


# ---------------------------------------------------------------------------
# Check 5: Missing TDS Detection
# ---------------------------------------------------------------------------

def detect_missing_tds(
    tally_data: dict,
    form26_entries: list[dict],
    matched_tally_keys: set[str],
) -> list[dict]:
    """Find Tally expense entries where TDS should have been deducted
    but no corresponding Form 26 entry exists.

    Strategy:
    1. Collect all TDS-applicable Tally expenses (by expense head)
    2. Group by vendor
    3. Check aggregate annual amounts against thresholds
    4. If above threshold and no Form 26 entry → flag as missing TDS
    """
    findings = []

    # Build set of Form 26 vendors (normalized) per section
    f26_vendors_by_section = defaultdict(set)
    for e in form26_entries:
        f26_vendors_by_section[e["section"]].add(normalize_name(e["vendor_name"]))

    # Collect TDS-applicable Tally expenses not already matched
    vendor_expenses = defaultdict(lambda: {
        "total_amount": 0,
        "entries": [],
        "expense_heads": set(),
        "vendor_name": "",
        "expected_sections": set(),
    })

    # Track journal entries by (vendor_norm, amount, date_month) to deduplicate
    # against GST Exp Register (same transaction appears in both registers)
    journal_seen = set()  # (vendor_norm, amount, date_month)

    # Process Journal Register entries
    for entry in tally_data.get("journal_register", {}).get("entries", []):
        key = _tally_entry_key("journal", entry)
        if key in matched_tally_keys:
            continue

        entry_type = entry.get("entry_type", "")
        if entry_type in ("tds_deduction", "salary", "discount", "other"):
            continue  # Not TDS-applicable expenses

        vendor = entry.get("particulars", "")
        if not vendor or not vendor.strip():
            continue  # Skip entries with no vendor/particulars

        vendor_norm = normalize_name(vendor)

        # Use individual posting amounts (not gross_total) as the expense amount
        # This avoids double-counting when gross_total includes non-expense heads
        postings = entry.get("account_postings", {})
        heads = [h for h in postings if h not in ("Gross Total", "Value")]

        date_str = str(entry.get("date", ""))[:7]  # YYYY-MM

        for head in heads:
            head_amount = abs(postings.get(head, 0))
            if head_amount == 0:
                continue

            # Record for dedup against GST Exp Register
            journal_seen.add((vendor_norm, round(head_amount, 0), date_str, head.lower()))

            sections = classify_expense_head(head)
            ve = vendor_expenses[(vendor_norm, tuple(sorted(sections)))]
            ve["total_amount"] += head_amount
            ve["entries"].append(entry)
            ve["expense_heads"].add(head)
            ve["vendor_name"] = vendor
            ve["expected_sections"].update(sections)

    # Track cross-register duplicates for flagging
    cross_register_dupes = []  # list of (vendor, amount, head, journal_date, gst_date)

    # Process GST Exp Register entries — flag duplicates with journal for review
    for entry in tally_data.get("purchase_gst_exp_register", {}).get("entries", []):
        key = _tally_entry_key("gst_exp", entry)
        if key in matched_tally_keys:
            continue

        vendor = entry.get("particulars", "")
        if not vendor or not vendor.strip():
            continue

        vendor_norm = normalize_name(vendor)
        base_amount = entry.get("base_amount", 0) or 0
        date_str = str(entry.get("date", ""))[:7]

        heads = list((entry.get("expense_heads") or {}).keys())

        # Check for duplicates: if journal already has this vendor + amount + month
        # for the same expense head, count once but flag for human review
        is_duped = False
        for head in heads:
            head_lower = head.lower().replace("_18%", "").replace("_12%", "").replace("_5%", "").strip()
            dedup_key = (vendor_norm, round(abs(base_amount), 0), date_str, head_lower)
            if dedup_key in journal_seen:
                is_duped = True
                cross_register_dupes.append({
                    "vendor": vendor,
                    "amount": base_amount,
                    "head": head,
                    "date": date_str,
                })
                break

        if is_duped:
            continue  # Counted from journal — will be flagged for review below

        for head in heads:
            sections = classify_expense_head(head)
            ve = vendor_expenses[(vendor_norm, tuple(sorted(sections)))]
            ve["total_amount"] += base_amount
            ve["entries"].append(entry)
            ve["expense_heads"].add(head)
            ve["vendor_name"] = vendor
            ve["expected_sections"].update(sections)

    # Now check each vendor group against thresholds
    for (vendor_norm, sections_tuple), ve in vendor_expenses.items():
        expected_sections = ve["expected_sections"] - {"unknown"}
        if not expected_sections:
            continue

        # Check if vendor has Form 26 entries in any expected section
        # Use both exact and fuzzy name matching
        vendor_in_f26 = False
        fuzzy_match_name = None
        for sec in expected_sections:
            f26_vendors = f26_vendors_by_section.get(sec, set())
            if vendor_norm in f26_vendors:
                vendor_in_f26 = True
                break
            # Fuzzy: check if any F26 vendor name is a substring or shares tokens
            vendor_tokens = set(vendor_norm.split())
            for f26_v in f26_vendors:
                f26_tokens = set(f26_v.split())
                overlap = vendor_tokens & f26_tokens
                # If >50% tokens overlap, consider it a likely match
                if overlap and len(overlap) >= min(len(vendor_tokens), len(f26_tokens)) * 0.5:
                    vendor_in_f26 = True
                    fuzzy_match_name = f26_v
                    break
            if vendor_in_f26:
                break

        if vendor_in_f26:
            # Vendor exists in Form 26, but unmatched Tally expenses may still
            # exceed threshold — don't skip, let it fall through to threshold check.
            # Flag as "partial coverage" instead of "missing TDS" if flagged.
            pass

        total = ve["total_amount"]
        if total <= 0:
            continue

        # Check against threshold for each expected section
        for sec in expected_sections:
            threshold = TDS_THRESHOLDS.get(sec, {}).get("aggregate_annual")
            if threshold and total < threshold:
                continue  # Below threshold — no TDS required

            # Collect review flags
            review_reasons = []

            # Flag 1: Year-end or quarter-end entries that don't fit in one go
            entry_dates = set()
            for ent in ve["entries"]:
                d = str(ent.get("date", ""))[:10]
                entry_dates.add(d)
            quarter_ends = {"03-31", "06-30", "09-30", "12-31"}
            qe_dates = [d for d in entry_dates if any(qe in d for qe in quarter_ends)]
            non_qe_dates = [d for d in entry_dates if not any(qe in d for qe in quarter_ends)]
            if qe_dates and non_qe_dates:
                review_reasons.append(
                    f"Has entries on quarter/year-end dates ({', '.join(sorted(qe_dates))}) "
                    f"alongside other dates — may include provisions or reversals."
                )

            # Flag 2: Cross-register duplicates for this vendor
            vendor_dupes = [d for d in cross_register_dupes
                            if normalize_name(d["vendor"]) == vendor_norm]
            if vendor_dupes:
                dupe_amt = vendor_dupes[0]["amount"]
                review_reasons.append(
                    f"Same transaction (₹{dupe_amt:,.0f}) appears in both "
                    f"Journal Register and GST Exp Register — counted once from "
                    f"Journal. Please confirm these are the same entry."
                )

            # Flag 3: Cr/Dr direction info (read but not interpreted)
            cr_entries = []
            dr_entries = []
            for ent in ve["entries"]:
                direction = ent.get("gross_total_direction", "")
                if direction == "cr":
                    cr_entries.append(ent)
                elif direction == "dr":
                    dr_entries.append(ent)
            if cr_entries and dr_entries:
                review_reasons.append(
                    f"Has both Dr ({len(dr_entries)}) and Cr ({len(cr_entries)}) "
                    f"entries — Cr entries may be reversals. Verify net expense."
                )

            needs_review = len(review_reasons) > 0
            severity = "error" if total >= (threshold or 0) else "warning"
            if needs_review:
                severity = "warning"

            review_note = ""
            if review_reasons:
                review_note = " Review: " + " | ".join(review_reasons)

            if vendor_in_f26:
                check_type = "partial_tds"
                msg = (f"Vendor {ve['vendor_name']} has Form 26 entry under {sec} "
                       f"but additional ₹{total:,.0f} in books "
                       f"({', '.join(sorted(ve['expense_heads']))}) "
                       f"is not covered. TDS may be short-deducted.{review_note}")
            else:
                check_type = "missing_tds"
                msg = (f"No Form 26 entry found for {ve['vendor_name']} "
                       f"under {sec}. Tally shows ₹{total:,.0f} in "
                       f"{', '.join(sorted(ve['expense_heads']))}. "
                       f"TDS may be missing.{review_note}")

            findings.append({
                "check": check_type,
                "severity": severity,
                "vendor": ve["vendor_name"],
                "vendor_normalized": vendor_norm,
                "expected_section": sec,
                "aggregate_amount": round(total, 2),
                "threshold": threshold,
                "num_tally_entries": len(ve["entries"]),
                "expense_heads": sorted(ve["expense_heads"]),
                "needs_review": needs_review,
                "review_reasons": review_reasons,
                "vendor_in_f26": vendor_in_f26,
                "message": msg,
            })

    return findings


def _tally_entry_key(source: str, entry: dict) -> str:
    """Create a unique key for a tally entry to track which are already matched."""
    return f"{source}|{entry.get('particulars','')}|{entry.get('voucher_no','')}|{entry.get('date','')}"


def build_matched_tally_keys(matches: list[dict]) -> set[str]:
    """Extract unique keys for all tally entries that were matched."""
    keys = set()
    for m in matches:
        for t in m.get("tally_entries", []):
            src = t.get("tally_source", "")
            # Map matcher source names back to register sources
            if src.startswith("journal"):
                src = "journal"
            elif src.startswith("gst_exp"):
                src = "gst_exp"
            else:
                src = src
            keys.add(f"{src}|{t.get('party_name','')}|{t.get('voucher_no','')}|{t.get('date','')}")
    return keys


# ---------------------------------------------------------------------------
# Check 6: TDS Timing Compliance (Deduction + Deposit)
# ---------------------------------------------------------------------------

def parse_challan_register(filepath: str) -> list[dict]:
    """Parse Form 26 Challan Register to extract deposit dates per section."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Challan Details"]

    challans = []
    current_section = None

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        sec = row[1].value
        if sec and "Total" not in str(sec) and "Grand" not in str(sec):
            current_section = str(sec).strip()

        challan_sr = str(row[2].value or "").strip()
        if not challan_sr or not current_section:
            continue

        deposit_date = row[12].value  # Cha/Vch Date
        if not deposit_date:
            continue

        # Parse quarter from "1 : Q2"
        quarter = ""
        if ":" in challan_sr:
            quarter = challan_sr.split(":")[-1].strip()

        challans.append({
            "section": current_section,
            "challan_sr": challan_sr,
            "quarter": quarter,
            "amount": row[10].value or 0,  # Total Paid
            "challan_no": row[11].value,
            "deposit_date": deposit_date,
        })

    wb.close()
    return challans


def get_deposit_due_date(deduction_date: datetime) -> datetime:
    """Calculate TDS deposit due date per Rule 30.
    - March deductions: due April 30
    - All other months: due 7th of next month
    """
    if deduction_date.month == 3:
        return datetime(deduction_date.year, 4, 30)
    elif deduction_date.month == 12:
        return datetime(deduction_date.year + 1, 1, 7)
    else:
        return datetime(deduction_date.year, deduction_date.month + 1, 7)


def find_challan_for_entry(challans: list[dict], section: str,
                           deduction_date: datetime) -> dict | None:
    """Find the challan that covers a specific TDS deduction.
    Match by section + closest deposit date on or after the deduction date.
    """
    section_challans = [c for c in challans if c["section"] == section]
    if not section_challans:
        return None

    # Find challan with deposit date closest to (and >= ) deduction date
    best = None
    best_diff = None
    for c in section_challans:
        dep_date = parse_date(c["deposit_date"])
        if not dep_date:
            continue
        diff = (dep_date - deduction_date).days
        if diff >= -7:  # Allow 7 days before (challan could slightly predate deduction)
            if best_diff is None or abs(diff) < abs(best_diff):
                best = c
                best_diff = diff

    return best


def months_between(d1: datetime, d2: datetime) -> int:
    """Calculate months (or part of month) between two dates for penalty.
    Returns at least 1 if any days late."""
    days = (d2 - d1).days
    if days <= 0:
        return 0
    return math.ceil(days / 30)


def check_tds_timing(matches: list[dict], challans: list[dict]) -> list[dict]:
    """Check TDS deduction and deposit timing for all matched entries.

    For each match (including each Tally entry in aggregated matches):
    - Compare expense date (Tally) vs deduction date (Form 26)
    - Compare deposit date (challan) vs deposit due date (Rule 30)

    Returns list of timing findings.
    """
    findings = []

    for m in matches:
        f26 = m["form26_entry"]
        tally_entries = m["tally_entries"]
        section = f26.get("section", "")
        tds_amount = f26.get("tax_deducted", 0) or 0
        deduction_date = parse_date(f26.get("tax_deducted_date"))

        if tds_amount == 0 or not deduction_date:
            continue

        # Per-entry deduction timing check
        for t in tally_entries:
            expense_date = parse_date(t.get("date"))
            if not expense_date:
                continue

            days_diff = (deduction_date - expense_date).days

            if days_diff > 0:
                # Late deduction: TDS deducted after expense date
                months_late = months_between(expense_date, deduction_date)
                # For aggregated matches, prorate TDS by entry amount
                entry_amount = t.get("amount", 0) or 0
                total_tally = sum(te.get("amount", 0) or 0 for te in tally_entries)
                if total_tally > 0 and len(tally_entries) > 1:
                    entry_tds = round(tds_amount * entry_amount / total_tally, 2)
                else:
                    entry_tds = tds_amount

                penalty = round(entry_tds * 0.01 * months_late, 2)

                findings.append({
                    "check": "tds_timing",
                    "sub_check": "late_deduction",
                    "severity": "error",
                    "vendor": f26.get("vendor_name", ""),
                    "pan": f26.get("pan", ""),
                    "form26_section": section,
                    "expense_date": expense_date.isoformat()[:10],
                    "deduction_date": deduction_date.isoformat()[:10],
                    "days_late": days_diff,
                    "months_late": months_late,
                    "expense_amount": entry_amount,
                    "tds_amount": entry_tds,
                    "penalty_rate_pct": 1.0,
                    "estimated_penalty": penalty,
                    "tally_party": t.get("party_name", ""),
                    "message": (
                        f"Late TDS deduction on {f26['vendor_name']} "
                        f"({section}): Expense on {expense_date.strftime('%d-%b-%Y')}, "
                        f"TDS deducted on {deduction_date.strftime('%d-%b-%Y')} "
                        f"({days_diff} days late, {months_late} month(s)). "
                        f"Interest u/s 201(1A): ₹{penalty:,.0f} "
                        f"({entry_tds:,.0f} × 1% × {months_late} months)."
                    ),
                })

            elif days_diff < -30:
                # Advance deduction — unusual, flag as info
                findings.append({
                    "check": "tds_timing",
                    "sub_check": "advance_deduction",
                    "severity": "info",
                    "vendor": f26.get("vendor_name", ""),
                    "form26_section": section,
                    "expense_date": expense_date.isoformat()[:10],
                    "deduction_date": deduction_date.isoformat()[:10],
                    "days_early": abs(days_diff),
                    "message": (
                        f"TDS deducted {abs(days_diff)} days before expense for "
                        f"{f26['vendor_name']} ({section}). Unusual — verify."
                    ),
                })

        # Deposit timing check
        if challans:
            challan = find_challan_for_entry(challans, section, deduction_date)
            deposit_due = get_deposit_due_date(deduction_date)

            if challan:
                actual_deposit = parse_date(challan["deposit_date"])
                if actual_deposit and actual_deposit > deposit_due:
                    days_late = (actual_deposit - deposit_due).days
                    months_late = months_between(deposit_due, actual_deposit)
                    penalty = round(tds_amount * 0.015 * months_late, 2)

                    findings.append({
                        "check": "tds_timing",
                        "sub_check": "late_deposit",
                        "severity": "error",
                        "vendor": f26.get("vendor_name", ""),
                        "pan": f26.get("pan", ""),
                        "form26_section": section,
                        "deduction_date": deduction_date.isoformat()[:10],
                        "deposit_due_date": deposit_due.isoformat()[:10],
                        "actual_deposit_date": actual_deposit.isoformat()[:10],
                        "days_late": days_late,
                        "months_late": months_late,
                        "tds_amount": tds_amount,
                        "penalty_rate_pct": 1.5,
                        "estimated_penalty": penalty,
                        "challan_no": challan.get("challan_no"),
                        "message": (
                            f"Late TDS deposit for {f26['vendor_name']} "
                            f"({section}): Due by {deposit_due.strftime('%d-%b-%Y')}, "
                            f"deposited on {actual_deposit.strftime('%d-%b-%Y')} "
                            f"({days_late} days late). "
                            f"Interest u/s 201(1A): ₹{penalty:,.0f} "
                            f"({tds_amount:,.0f} × 1.5% × {months_late} months)."
                        ),
                    })

    return findings


# ---------------------------------------------------------------------------
# Main Runner
# ---------------------------------------------------------------------------

def run(parsed_dir: str, results_dir: str, event_callback=None) -> dict:
    """Run all TDS compliance checks.

    Args:
        parsed_dir: Path to parsed data (parsed_form26.json, parsed_tally.json)
        results_dir: Path to results (match_results.json, output checker_results.json)
        event_callback: Optional callable(agent, message, type) for real-time progress.
    """

    def emit(message, type="detail"):
        if event_callback:
            event_callback("TDS Checker", message, type)
        print(f"  [{type}] {message}")

    parsed_path = Path(parsed_dir)
    results_path = Path(results_dir)

    # Load data
    with open(parsed_path / "parsed_form26.json") as f:
        form26_data = json.load(f)
    with open(parsed_path / "parsed_tally.json") as f:
        tally_data = json.load(f)
    with open(results_path / "match_results.json") as f:
        match_data = json.load(f)

    form26_entries = form26_data["entries"]
    matches = match_data["matches"]

    all_findings = []

    # ---- Section Validation ----
    emit(f"Validating TDS sections for {len(matches)} matched entries...")
    section_findings = []
    for m in matches:
        finding = check_section(m)
        if finding:
            section_findings.append(finding)
    all_findings.extend(section_findings)
    if section_findings:
        emit(f"{len(section_findings)} section classification issues found", "warning")
    else:
        emit("All sections correctly classified", "success")

    # ---- Rate + PAN Validation ----
    emit("Validating TDS rates and PAN compliance...")
    rate_findings = []
    pan_findings = []
    for m in matches:
        results = check_rate(m)
        for f in results:
            if f["check"] == "pan_validation":
                pan_findings.append(f)
            else:
                rate_findings.append(f)
    all_findings.extend(rate_findings)
    all_findings.extend(pan_findings)
    if pan_findings:
        emit(f"{len(pan_findings)} PAN issue(s) — Section 206AA may apply", "warning")
    if rate_findings:
        emit(f"{len(rate_findings)} rate mismatch(es) found", "warning")
    if not rate_findings and not pan_findings:
        emit("All TDS rates and PAN compliance verified", "success")

    # ---- Base Amount Validation ----
    emit("Checking if TDS computed on correct base amount (pre-GST)...")
    base_findings = []
    for m in matches:
        finding = check_base_amount(m)
        if finding:
            base_findings.append(finding)
    all_findings.extend(base_findings)
    if base_findings:
        emit(f"{len(base_findings)} base amount issues — TDS may be on GST-inclusive amount", "warning")
    else:
        emit("All base amounts correct", "success")

    # ---- Threshold Validation ----
    emit("Checking threshold compliance across vendors...")
    threshold_findings = check_thresholds(matches, tally_data)
    all_findings.extend(threshold_findings)
    below_count = sum(1 for f in threshold_findings if "below" in f.get("status", ""))
    if below_count:
        emit(f"{below_count} vendors below annual threshold (TDS not mandatory)")

    # ---- Missing TDS Detection ----
    # Count TDS-applicable entries by expense type for natural progress
    jr_entries = tally_data.get("journal_register", {}).get("entries", [])
    gst_entries = tally_data.get("purchase_gst_exp_register", {}).get("entries", [])
    interest_count = sum(1 for e in jr_entries if e.get("entry_type") == "interest_payment")
    freight_count = sum(1 for e in jr_entries if e.get("entry_type") == "freight_expense")
    brokerage_count = sum(1 for e in jr_entries if e.get("entry_type") == "brokerage")
    professional_count = sum(1 for e in jr_entries if e.get("entry_type") in ("professional_fees", "consultancy", "audit_fees"))
    emit("Scanning books for entries where TDS may be missing...")
    if interest_count:
        emit(f"Scanning interest expenses... {interest_count} entries")
    if freight_count:
        emit(f"Scanning contractor/freight expenses... {freight_count} entries")
    if brokerage_count:
        emit(f"Scanning brokerage/commission expenses... {brokerage_count} entries")
    if professional_count:
        emit(f"Scanning professional/consultancy fees... {professional_count} entries")
    if gst_entries:
        emit(f"Scanning GST expense register... {len(gst_entries)} entries")
    matched_keys = build_matched_tally_keys(matches)
    missing_findings = detect_missing_tds(tally_data, form26_entries, matched_keys)
    all_findings.extend(missing_findings)
    review_count = sum(1 for f in missing_findings if f.get("needs_review"))
    error_count = sum(1 for f in missing_findings if f["severity"] == "error")
    if missing_findings:
        parts = []
        if error_count:
            parts.append(f"{error_count} missing TDS")
        if review_count:
            parts.append(f"{review_count} flagged for review")
        emit(f"Found {', '.join(parts)}", "warning")
    else:
        emit("No missing TDS detected", "success")

    # ---- TDS Timing Compliance ----
    emit("Checking TDS deduction and deposit timing...")
    # Try to load challan register
    challans = []
    base_dir = Path(__file__).parent.parent.parent  # repo root
    challan_paths = [
        base_dir / "data" / "hpc" / "Form 26 challan register.xlsx",
        parsed_path.parent / "uploads" / "challan.xlsx",
        parsed_path / "challan_register.json",
    ]
    for cp in challan_paths:
        if cp.exists() and cp.suffix == ".xlsx":
            try:
                challans = parse_challan_register(str(cp))
                emit(f"Loaded {len(challans)} challan entries from {cp.name}")
            except Exception as e:
                emit(f"Could not parse challan register: {e}", "warning")
            break

    timing_findings = check_tds_timing(matches, challans)
    all_findings.extend(timing_findings)

    late_deductions = [f for f in timing_findings if f.get("sub_check") == "late_deduction"]
    late_deposits = [f for f in timing_findings if f.get("sub_check") == "late_deposit"]
    total_penalty = sum(f.get("estimated_penalty", 0) for f in timing_findings)

    if late_deductions:
        emit(f"{len(late_deductions)} late TDS deduction(s) found — estimated interest ₹{sum(f['estimated_penalty'] for f in late_deductions):,.0f}", "warning")
    if late_deposits:
        emit(f"{len(late_deposits)} late TDS deposit(s) found — estimated interest ₹{sum(f['estimated_penalty'] for f in late_deposits):,.0f}", "warning")
    if not late_deductions and not late_deposits:
        emit("All TDS deducted and deposited on time", "success")
    if not challans:
        emit("Challan register not found — deposit timing not checked", "warning")

    # ---- Summarize ----
    summary = {
        "total_findings": len(all_findings),
        "by_check": {
            "section_validation": len(section_findings),
            "rate_validation": len(rate_findings),
            "pan_validation": len(pan_findings),
            "base_amount_validation": len(base_findings),
            "threshold_validation": len(threshold_findings),
            "missing_tds": len(missing_findings),
            "tds_timing": len(timing_findings),
        },
        "by_severity": {
            "error": sum(1 for f in all_findings if f["severity"] == "error"),
            "warning": sum(1 for f in all_findings if f["severity"] == "warning"),
            "info": sum(1 for f in all_findings if f["severity"] == "info"),
        },
        "matches_checked": len(matches),
        "sections_in_scope": sorted(set(m["form26_entry"]["section"] for m in matches)),
    }

    results = {
        "run_timestamp": datetime.now().isoformat(),
        "summary": summary,
        "findings": all_findings,
    }

    # Write output
    out_file = results_path / "checker_results.json"
    with open(out_file, "w") as f:
        json.dump(results, f, indent=2, default=str)
    print(f"\n[TDS Checker] Wrote {out_file}")

    # Print summary
    print("\n" + "=" * 60)
    print("TDS CHECKER AGENT — SUMMARY")
    print("=" * 60)
    print(f"\nMatches checked: {summary['matches_checked']}")
    print(f"Sections in scope: {', '.join(summary['sections_in_scope'])}")
    print(f"\nTotal findings: {summary['total_findings']}")
    print(f"  Errors:   {summary['by_severity']['error']}")
    print(f"  Warnings: {summary['by_severity']['warning']}")
    print(f"  Info:     {summary['by_severity']['info']}")
    print(f"\nBy check:")
    for check, count in summary["by_check"].items():
        print(f"  {check}: {count}")

    if all_findings:
        print(f"\n{'─' * 60}")
        print("TOP FINDINGS:")
        print(f"{'─' * 60}")
        # Show errors first, then warnings
        for f in sorted(all_findings, key=lambda x: {"error": 0, "warning": 1, "info": 2}[x["severity"]]):
            icon = {"error": "✗", "warning": "⚠", "info": "ℹ"}[f["severity"]]
            print(f"\n  {icon} [{f['severity'].upper()}] {f['check']}")
            print(f"    {f['message']}")

    return results


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    base = Path(__file__).parent.parent
    parsed = base / "data" / "parsed"
    output = base / "data" / "results"

    run(str(parsed), str(output))
