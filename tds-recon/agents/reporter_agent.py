"""
Reporter Agent — TDS Reconciliation MVP
========================================
Consolidates match results + compliance findings into actionable outputs
that a CA/accountant can use for Form 26 filing validation.

Generates:
  1. reconciliation_summary.json  — Executive summary + metrics
  2. reconciliation_report.csv    — Full match details (flat table)
  3. findings_report.csv          — All findings with remediation guidance

Inputs:
  - data/parsed/parsed_form26.json
  - data/parsed/parsed_tally.json
  - data/results/match_results.json
  - data/results/checker_results.json
"""

import csv
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path


# ---------------------------------------------------------------------------
# Remediation guidance per finding type
# ---------------------------------------------------------------------------

REMEDIATION = {
    "section_validation": {
        "mismatch": (
            "Review the vendor invoice/contract to confirm the nature of service. "
            "If the service is professional/technical, amend Form 26 to the correct "
            "section before quarterly filing deadline."
        ),
        "review": (
            "Ambiguous classification — verify invoice nature. If advertisement is "
            "creative/professional service → 194J(b). If it's production/printing "
            "contract work → 194C. Obtain vendor declaration if uncertain."
        ),
        "unclassified": (
            "Expense head could not be auto-classified. Manually review the vendor "
            "invoice and assign the correct TDS section."
        ),
    },
    "rate_validation": {
        "default": (
            "TDS rate does not match expected rate for this section + entity type. "
            "Verify the vendor's PAN status and entity classification. "
            "If rate is wrong, compute differential TDS and deposit with interest u/s 201(1A)."
        ),
    },
    "base_amount_validation": {
        "default": (
            "TDS appears computed on GST-inclusive amount instead of base amount. "
            "As per CBDT Circular 23/2017, TDS should be on the amount excluding GST "
            "when GST is shown separately. Recalculate and claim refund of excess TDS "
            "or adjust in next quarter's challan."
        ),
    },
    "threshold_validation": {
        "below_threshold_but_deducted": (
            "TDS was deducted even though aggregate amount is below the threshold. "
            "This is not an error — voluntary TDS deduction is valid. "
            "No action needed, but note for internal records."
        ),
    },
    "missing_tds": {
        "default": (
            "Tally shows expenses to this vendor but no TDS was deducted in Form 26. "
            "Action required: (1) Check if vendor provided Form 15G/15H or lower "
            "deduction certificate. (2) If not, deduct TDS immediately and deposit "
            "with interest u/s 201(1A) at 1% per month. (3) File revised TDS return "
            "for the applicable quarter."
        ),
    },
}


def get_remediation(check: str, status: str = "default") -> str:
    """Get remediation guidance for a finding."""
    check_remediation = REMEDIATION.get(check, {})
    return check_remediation.get(status, check_remediation.get("default", "Manual review required."))


# ---------------------------------------------------------------------------
# Report Builders
# ---------------------------------------------------------------------------

def build_executive_summary(
    form26_data: dict,
    match_data: dict,
    checker_data: dict,
) -> dict:
    """Build executive summary with key metrics."""
    matches = match_data["matches"]
    findings = checker_data["findings"]
    f26_entries = form26_data["entries"]

    # Section-wise breakdown
    section_stats = defaultdict(lambda: {
        "form26_count": 0, "form26_amount": 0, "form26_tds": 0,
        "matched_count": 0, "matched_amount": 0, "matched_tds": 0,
    })

    for e in f26_entries:
        sec = e["section"]
        section_stats[sec]["form26_count"] += 1
        section_stats[sec]["form26_amount"] += e.get("amount_paid", 0)
        section_stats[sec]["form26_tds"] += e.get("tax_deducted", 0)

    for m in matches:
        sec = m["form26_entry"]["section"]
        section_stats[sec]["matched_count"] += 1
        section_stats[sec]["matched_amount"] += m["form26_entry"].get("amount_paid", 0)
        section_stats[sec]["matched_tds"] += m["form26_entry"].get("tax_deducted", 0)

    # Match quality
    confidences = [m.get("confidence", 1.0) for m in matches]
    pass_distribution = defaultdict(int)
    for m in matches:
        pass_distribution[m.get("pass_name", f"pass{m['pass']}")] += 1

    # Findings severity
    severity_counts = defaultdict(int)
    check_counts = defaultdict(int)
    for f in findings:
        severity_counts[f["severity"]] += 1
        check_counts[f["check"]] += 1

    # Revenue impact of findings
    missing_tds_amount = sum(
        f.get("aggregate_amount", 0) for f in findings if f["check"] == "missing_tds"
    )

    summary = match_data.get("summary", {})

    return {
        "report_generated": datetime.now().isoformat(),
        "assessment_year": "2025-26",
        "financial_year": "2024-25",
        "scope": {
            "sections_reconciled": summary.get("sections_processed",
                                               sorted(set(m["form26_entry"]["section"] for m in matches))),
            "sections_pending": sorted(
                set(e["section"] for e in f26_entries)
                - set(m["form26_entry"]["section"] for m in matches)
            ),
        },
        "matching": {
            "form26_total": len(f26_entries),
            "form26_in_scope": summary.get("form26_total", len(matches)),
            "matched_with_tds": summary.get("form26_matched", len(matches)),
            "below_threshold_resolved": summary.get("below_threshold_resolved", 0),
            "total_resolved": summary.get("total_resolved", summary.get("form26_matched", len(matches))),
            "unmatched": summary.get("form26_unmatched", 0),
            "match_rate_pct": round(
                summary.get("total_resolved", summary.get("form26_matched", len(matches)))
                / max(summary.get("total_resolved", 1) + summary.get("form26_unmatched", 0), 1) * 100, 1
            ),
            "avg_confidence": round(sum(confidences) / max(len(confidences), 1), 3),
            "by_pass": dict(pass_distribution),
        },
        "section_wise": {
            sec: stats for sec, stats in sorted(section_stats.items())
        },
        "compliance": {
            "total_findings": len(findings),
            "errors": severity_counts.get("error", 0),
            "warnings": severity_counts.get("warning", 0),
            "info": severity_counts.get("info", 0),
            "by_check": dict(check_counts),
            "missing_tds_exposure": round(missing_tds_amount, 2),
            "clean_bill": len(findings) == 0 or severity_counts.get("error", 0) == 0,
        },
        "amounts": {
            "total_form26_payments": sum(e.get("amount_paid", 0) for e in f26_entries),
            "total_form26_tds": sum(e.get("tax_deducted", 0) for e in f26_entries),
            "matched_payments": sum(m["form26_entry"].get("amount_paid", 0) for m in matches),
            "matched_tds": sum(m["form26_entry"].get("tax_deducted", 0) for m in matches),
        },
    }


def build_match_report_rows(matches: list[dict]) -> list[dict]:
    """Build flat CSV rows for the reconciliation report."""
    rows = []
    for i, m in enumerate(matches, 1):
        f26 = m["form26_entry"]
        tally_entries = m["tally_entries"]

        # Collect expense categories
        expense_cats = set()
        for t in tally_entries:
            if t.get("expense_heads"):
                expense_cats.update(t["expense_heads"].keys())
            elif t.get("account_postings"):
                expense_cats.update(
                    k for k in t["account_postings"]
                    if k not in ("Gross Total", "Value")
                )
            src = t.get("tally_source", "")
            if src == "journal_interest":
                expense_cats.add("Interest Paid")
            elif src == "journal_freight":
                expense_cats.add("Freight Charges")

        tally_amount = sum(t.get("amount", 0) for t in tally_entries)
        tally_party = tally_entries[0].get("party_name", "") if tally_entries else ""
        tally_date = tally_entries[0].get("date", "")[:10] if tally_entries else ""

        rows.append({
            "sr_no": i,
            "vendor_name": f26.get("vendor_name", ""),
            "pan": f26.get("pan", ""),
            "section": f26.get("section", ""),
            "form26_amount": f26.get("amount_paid", 0),
            "form26_date": str(f26.get("amount_paid_date", ""))[:10],
            "tds_rate_pct": f26.get("tax_rate_pct", ""),
            "tds_amount": f26.get("tax_deducted", 0),
            "tally_party": tally_party,
            "tally_amount": round(tally_amount, 2),
            "tally_date": tally_date,
            "tally_entries_count": len(tally_entries),
            "amount_diff": round(f26.get("amount_paid", 0) - tally_amount, 2),
            "match_type": m.get("pass_name", ""),
            "confidence": m.get("confidence", 1.0),
            "expense_category": "; ".join(sorted(expense_cats)),
            "status": "Matched",
        })

    rows.sort(key=lambda x: (x["section"], x["vendor_name"], x["form26_date"]))
    for i, r in enumerate(rows, 1):
        r["sr_no"] = i
    return rows


def build_findings_report_rows(findings: list[dict]) -> list[dict]:
    """Build flat CSV rows for the findings report with remediation."""
    rows = []
    for i, f in enumerate(findings, 1):
        status = f.get("status", "default")
        remediation = get_remediation(f["check"], status)

        rows.append({
            "sr_no": i,
            "severity": f["severity"].upper(),
            "check_type": f["check"],
            "vendor": f.get("vendor", ""),
            "pan": f.get("pan", ""),
            "section": f.get("form26_section", f.get("expected_section", "")),
            "amount": f.get("form26_amount", f.get("aggregate_amount", "")),
            "finding": f["message"],
            "remediation": remediation,
            "status": "Open",
        })

    # Sort: errors first, then warnings, then info
    severity_order = {"ERROR": 0, "WARNING": 1, "INFO": 2}
    rows.sort(key=lambda x: severity_order.get(x["severity"], 9))
    for i, r in enumerate(rows, 1):
        r["sr_no"] = i
    return rows


# ---------------------------------------------------------------------------
# 4b. Excel Report (3 sheets)
# ---------------------------------------------------------------------------

def build_excel_report(
    filepath,
    finding_rows: list[dict],
    match_rows: list[dict],
    below_threshold_entries: list[dict],
    exemptions: list[dict],
    timing_findings: list[dict] = None,
):
    """Build a multi-sheet Excel workbook.

    Sheet 1: Issues for Human Review (findings — wrong section, missing TDS, etc.)
    Sheet 2: TDS Report — all matched entries with expense head
    Sheet 3: Zero TDS Entries — below-threshold and exempt with reason
    Sheet 4: Late TDS Deduction — per-entry deduction timing violations
    Sheet 5: Late TDS Deposit — challan deposit timing violations
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="467273", end_color="467273", fill_type="solid")
    error_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    warn_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    def write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    def style_row(ws, row_num, num_cols, fill=None):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    # ========== Sheet 1: Issues for Human Review ==========
    ws1 = wb.active
    ws1.title = "Issues for Review"
    h1 = ["Sr.", "Severity", "Check Type", "Vendor", "PAN", "Section",
          "Amount", "Finding", "Remediation", "Status"]
    write_header(ws1, h1)

    for i, row in enumerate(finding_rows, 2):
        ws1.cell(row=i, column=1, value=row.get("sr_no", i - 1))
        ws1.cell(row=i, column=2, value=row.get("severity", ""))
        ws1.cell(row=i, column=3, value=row.get("check_type", ""))
        ws1.cell(row=i, column=4, value=row.get("vendor", ""))
        ws1.cell(row=i, column=5, value=row.get("pan", ""))
        ws1.cell(row=i, column=6, value=row.get("section", ""))
        ws1.cell(row=i, column=7, value=row.get("amount", 0))
        ws1.cell(row=i, column=8, value=row.get("finding", ""))
        ws1.cell(row=i, column=9, value=row.get("remediation", ""))
        ws1.cell(row=i, column=10, value="Open")
        fill = error_fill if row.get("severity") == "ERROR" else warn_fill if row.get("severity") == "WARNING" else None
        style_row(ws1, i, len(h1), fill)

    for col in range(1, len(h1) + 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(12, min(40, len(str(h1[col - 1])) + 4))
    ws1.column_dimensions["H"].width = 60
    ws1.column_dimensions["I"].width = 60
    ws1.auto_filter.ref = ws1.dimensions

    # ========== Sheet 2: TDS Report (Matched Entries) ==========
    ws2 = wb.create_sheet("TDS Report - Matched")
    h2 = ["Sr.", "Vendor", "PAN", "Section", "Form 26 Amount", "Form 26 Date",
          "TDS Rate %", "TDS Amount", "Tally Party", "Tally Amount", "Tally Date",
          "Tally Entries", "Amount Diff", "Match Type", "Confidence", "Status", "Expense Head"]
    write_header(ws2, h2)

    for i, row in enumerate(match_rows, 2):
        ws2.cell(row=i, column=1, value=row.get("sr_no", i - 1))
        ws2.cell(row=i, column=2, value=row.get("vendor_name", ""))
        ws2.cell(row=i, column=3, value=row.get("pan", ""))
        ws2.cell(row=i, column=4, value=row.get("section", ""))
        ws2.cell(row=i, column=5, value=row.get("form26_amount", 0))
        ws2.cell(row=i, column=6, value=row.get("form26_date", ""))
        ws2.cell(row=i, column=7, value=row.get("tds_rate_pct", ""))
        ws2.cell(row=i, column=8, value=row.get("tds_amount", 0))
        ws2.cell(row=i, column=9, value=row.get("tally_party", ""))
        ws2.cell(row=i, column=10, value=row.get("tally_amount", 0))
        ws2.cell(row=i, column=11, value=row.get("tally_date", ""))
        ws2.cell(row=i, column=12, value=row.get("tally_entries_count", 0))
        ws2.cell(row=i, column=13, value=row.get("amount_diff", 0))
        ws2.cell(row=i, column=14, value=row.get("match_type", ""))
        ws2.cell(row=i, column=15, value=row.get("confidence", 0))
        ws2.cell(row=i, column=16, value="Reconciled")
        ws2.cell(row=i, column=17, value=row.get("expense_category", ""))
        style_row(ws2, i, len(h2))

    for col in range(1, len(h2) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(10, min(30, len(str(h2[col - 1])) + 4))
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["Q"].width = 40
    ws2.auto_filter.ref = ws2.dimensions

    # ========== Sheet 3: Zero TDS Entries ==========
    ws3 = wb.create_sheet("Zero TDS - Exempt")
    h3 = ["Sr.", "Vendor", "Section", "Amount", "Entries", "Reason for Zero TDS", "Expense Head"]
    write_header(ws3, h3)

    row_num = 2

    # Section threshold descriptions
    THRESHOLD_DESC = {
        "192": "₹2,50,000 basic exemption",
        "194A": "₹5,000 annual limit",
        "194C": "₹1,00,000 annual limit",
        "194H": "₹15,000 annual limit",
        "194J(a)": "₹30,000 annual limit",
        "194J(b)": "₹30,000 annual limit",
        "194Q": "₹50,00,000 annual limit",
    }

    # Below-threshold grouped by (vendor, section)
    vendor_groups = {}
    for e in below_threshold_entries:
        v = e.get("party_name", "Unknown")
        sec = e.get("_section", "194C")  # section tagged during collection
        key = (v, sec)
        if key not in vendor_groups:
            vendor_groups[key] = {"amount": 0, "count": 0, "heads": set(), "section": sec}
        vendor_groups[key]["amount"] += e.get("amount", 0)
        vendor_groups[key]["count"] += 1
        if e.get("expense_heads"):
            vendor_groups[key]["heads"].update(e["expense_heads"].keys())
        src = e.get("tally_source", "")
        if "freight" in src:
            vendor_groups[key]["heads"].add("Freight Charges")
        if "brokerage" in src:
            vendor_groups[key]["heads"].add("Brokerage and Commission")
        if "interest" in src:
            vendor_groups[key]["heads"].add("Interest Paid")

    for (v, sec), data in sorted(vendor_groups.items(), key=lambda x: -x[1]["amount"]):
        threshold_text = THRESHOLD_DESC.get(sec, "annual limit")
        ws3.cell(row=row_num, column=1, value=row_num - 1)
        ws3.cell(row=row_num, column=2, value=v)
        ws3.cell(row=row_num, column=3, value=sec)
        ws3.cell(row=row_num, column=4, value=round(data["amount"], 2))
        ws3.cell(row=row_num, column=5, value=data["count"])
        ws3.cell(row=row_num, column=6, value=f"Below threshold - aggregate Rs {data['amount']:,.0f} < {threshold_text}")
        ws3.cell(row=row_num, column=7, value="; ".join(sorted(data["heads"])) if data["heads"] else "")
        style_row(ws3, row_num, len(h3))
        row_num += 1

    # Exempt entries
    for e in exemptions:
        f26 = e.get("form26_entry", {})
        ws3.cell(row=row_num, column=1, value=row_num - 1)
        ws3.cell(row=row_num, column=2, value=f26.get("vendor_name", ""))
        ws3.cell(row=row_num, column=3, value=f26.get("section", ""))
        ws3.cell(row=row_num, column=4, value=f26.get("amount_paid", 0))
        ws3.cell(row=row_num, column=5, value=1)
        ws3.cell(row=row_num, column=6, value=f"Exempt - {e.get('pass_name', 'exempt filter')}")
        ws3.cell(row=row_num, column=7, value="")
        style_row(ws3, row_num, len(h3))
        row_num += 1

    for col in range(1, len(h3) + 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(10, min(30, len(str(h3[col - 1])) + 4))
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["F"].width = 50
    ws3.column_dimensions["G"].width = 40
    ws3.auto_filter.ref = ws3.dimensions

    # ========== Sheet 4: Late TDS Deduction ==========
    if timing_findings:
        late_deductions = [f for f in timing_findings if f.get("sub_check") == "late_deduction"]
        if late_deductions:
            ws4 = wb.create_sheet("Late TDS Deduction")
            h4 = ["Sr.", "Vendor", "PAN", "Section", "Expense Head",
                   "Expense Amount", "TDS Amount",
                   "Expense Date", "Deduction Date", "Days Late", "Months",
                   "Penalty Rate", "Estimated Interest", "Tally Party"]
            write_header(ws4, h4)

            for i, f in enumerate(late_deductions, 2):
                ws4.cell(row=i, column=1, value=i - 1)
                ws4.cell(row=i, column=2, value=f.get("vendor", ""))
                ws4.cell(row=i, column=3, value=f.get("pan", ""))
                ws4.cell(row=i, column=4, value=f.get("form26_section", ""))
                ws4.cell(row=i, column=5, value=f.get("tally_party", ""))
                ws4.cell(row=i, column=6, value=f.get("expense_amount", 0))
                ws4.cell(row=i, column=7, value=f.get("tds_amount", 0))
                ws4.cell(row=i, column=8, value=f.get("expense_date", ""))
                ws4.cell(row=i, column=9, value=f.get("deduction_date", ""))
                ws4.cell(row=i, column=10, value=f.get("days_late", 0))
                ws4.cell(row=i, column=11, value=f.get("months_late", 0))
                ws4.cell(row=i, column=12, value="1% per month")
                ws4.cell(row=i, column=13, value=f.get("estimated_penalty", 0))
                ws4.cell(row=i, column=14, value=f.get("tally_party", ""))
                fill = error_fill if f.get("days_late", 0) > 30 else warn_fill
                style_row(ws4, i, len(h4), fill)

            for col in range(1, len(h4) + 1):
                ws4.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(10, min(30, len(str(h4[col - 1])) + 4))
            ws4.column_dimensions["B"].width = 30
            ws4.column_dimensions["N"].width = 30
            ws4.auto_filter.ref = ws4.dimensions

        # ========== Sheet 5: Late TDS Deposit ==========
        late_deposits = [f for f in timing_findings if f.get("sub_check") == "late_deposit"]
        if late_deposits:
            ws5 = wb.create_sheet("Late TDS Deposit")
            h5 = ["Sr.", "Vendor", "PAN", "Section",
                   "TDS Amount", "Deduction Date",
                   "Deposit Due Date", "Actual Deposit Date", "Days Late", "Months",
                   "Penalty Rate", "Estimated Interest", "Challan No."]
            write_header(ws5, h5)

            for i, f in enumerate(late_deposits, 2):
                ws5.cell(row=i, column=1, value=i - 1)
                ws5.cell(row=i, column=2, value=f.get("vendor", ""))
                ws5.cell(row=i, column=3, value=f.get("pan", ""))
                ws5.cell(row=i, column=4, value=f.get("form26_section", ""))
                ws5.cell(row=i, column=5, value=f.get("tds_amount", 0))
                ws5.cell(row=i, column=6, value=f.get("deduction_date", ""))
                ws5.cell(row=i, column=7, value=f.get("deposit_due_date", ""))
                ws5.cell(row=i, column=8, value=f.get("actual_deposit_date", ""))
                ws5.cell(row=i, column=9, value=f.get("days_late", 0))
                ws5.cell(row=i, column=10, value=f.get("months_late", 0))
                ws5.cell(row=i, column=11, value="1.5% per month")
                ws5.cell(row=i, column=12, value=f.get("estimated_penalty", 0))
                ws5.cell(row=i, column=13, value=f.get("challan_no", ""))
                fill = error_fill if f.get("days_late", 0) > 30 else warn_fill
                style_row(ws5, i, len(h5), fill)

            for col in range(1, len(h5) + 1):
                ws5.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(10, min(30, len(str(h5[col - 1])) + 4))
            ws5.column_dimensions["B"].width = 30
            ws5.auto_filter.ref = ws5.dimensions

    wb.save(str(filepath))
    wb.close()


# ---------------------------------------------------------------------------
# Main Runner
# ---------------------------------------------------------------------------

def run(parsed_dir: str, results_dir: str) -> dict:
    """Generate all reports.

    Args:
        parsed_dir: Path to parsed data
        results_dir: Path to results (reads match + checker, writes reports)
    """
    parsed_path = Path(parsed_dir)
    results_path = Path(results_dir)

    # Load data
    with open(parsed_path / "parsed_form26.json") as f:
        form26_data = json.load(f)
    with open(parsed_path / "parsed_tally.json") as f:
        tally_data = json.load(f)
    with open(results_path / "match_results.json") as f:
        match_data = json.load(f)
    with open(results_path / "checker_results.json") as f:
        checker_data = json.load(f)

    matches = match_data["matches"]
    findings = checker_data["findings"]

    print("[Reporter] Generating reports...")

    # ---- 1. Executive Summary ----
    summary = build_executive_summary(form26_data, match_data, checker_data)
    summary_file = results_path / "reconciliation_summary.json"
    with open(summary_file, "w") as f:
        json.dump(summary, f, indent=2, default=str)
    print(f"  → {summary_file}")

    # ---- 2. Reconciliation Report (CSV) ----
    match_rows = build_match_report_rows(matches)
    report_file = results_path / "reconciliation_report.csv"
    if match_rows:
        fieldnames = list(match_rows[0].keys())
        with open(report_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(match_rows)
    print(f"  → {report_file} ({len(match_rows)} rows)")

    # ---- 3. Findings Report (CSV) ----
    finding_rows = build_findings_report_rows(findings)
    findings_file = results_path / "findings_report.csv"
    if finding_rows:
        fieldnames = list(finding_rows[0].keys())
        with open(findings_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(finding_rows)
    print(f"  → {findings_file} ({len(finding_rows)} rows)")

    # ---- 4. Excel Workbook (3 sheets) ----
    excel_file = results_path / "tds_recon_report.xlsx"
    # Collect below-threshold entries from ALL sections via unmatched_tally dict
    below_threshold = []
    unmatched_tally = match_data.get("unmatched_tally", {})
    for sec, entries in unmatched_tally.items():
        for e in entries:
            if e.get("_below_threshold"):
                below_threshold.append({**e, "_section": sec})
    # Backward compat: also check old keys
    for old_key, sec in [("unmatched_tally_194c", "194C"), ("unmatched_tally_194a", "194A")]:
        if not unmatched_tally and old_key in match_data:
            for e in match_data[old_key]:
                if e.get("_below_threshold"):
                    below_threshold.append({**e, "_section": sec})
    timing_findings = [f for f in findings if f.get("check") == "tds_timing"]
    build_excel_report(excel_file, finding_rows, match_rows, below_threshold,
                       match_data.get("exemptions", []), timing_findings=timing_findings)
    sheet_count = 3 + (1 if any(f.get("sub_check") == "late_deduction" for f in timing_findings) else 0) \
                    + (1 if any(f.get("sub_check") == "late_deposit" for f in timing_findings) else 0)
    print(f"  → {excel_file} ({sheet_count} sheets)")

    # ---- Print Summary ----
    print("\n" + "=" * 60)
    print("REPORTER AGENT — RECONCILIATION SUMMARY")
    print("=" * 60)
    print(f"\nAssessment Year: {summary['assessment_year']}")
    print(f"Sections Reconciled: {', '.join(summary['scope']['sections_reconciled'])}")
    print(f"Sections Pending: {', '.join(summary['scope']['sections_pending']) or 'None'}")

    print(f"\n--- MATCHING ---")
    m = summary["matching"]
    print(f"Form 26 entries (all):    {m['form26_total']}")
    print(f"Form 26 entries (scope):  {m['form26_in_scope']}")
    print(f"Matched (with TDS):       {m['matched_with_tds']}")
    print(f"Below threshold (TDS=0):  {m['below_threshold_resolved']}")
    print(f"Total resolved:           {m['total_resolved']} ({m['match_rate_pct']}%)")
    print(f"Unmatched:                {m['unmatched']}")
    print(f"Avg confidence:           {m['avg_confidence']}")
    print(f"By pass: {dict(m['by_pass'])}")

    print(f"\n--- SECTION-WISE ---")
    for sec, stats in sorted(summary["section_wise"].items()):
        in_scope = sec in summary["scope"]["sections_reconciled"]
        tag = "" if in_scope else " [NOT IN SCOPE]"
        print(f"  {sec}{tag}:")
        print(f"    Form 26: {stats['form26_count']} entries, "
              f"₹{stats['form26_amount']:,} paid, ₹{stats['form26_tds']:,} TDS")
        if in_scope:
            print(f"    Matched: {stats['matched_count']} entries, "
                  f"₹{stats['matched_amount']:,} paid, ₹{stats['matched_tds']:,} TDS")

    print(f"\n--- AMOUNTS ---")
    a = summary["amounts"]
    print(f"Total Form 26 payments: ₹{a['total_form26_payments']:,}")
    print(f"Total Form 26 TDS:      ₹{a['total_form26_tds']:,}")
    print(f"Matched payments:       ₹{a['matched_payments']:,}")
    print(f"Matched TDS:            ₹{a['matched_tds']:,}")

    print(f"\n--- COMPLIANCE ---")
    c = summary["compliance"]
    print(f"Total findings:    {c['total_findings']}")
    print(f"  Errors:          {c['errors']}")
    print(f"  Warnings:        {c['warnings']}")
    print(f"  Info:             {c['info']}")
    if c["missing_tds_exposure"] > 0:
        print(f"Missing TDS exposure: ₹{c['missing_tds_exposure']:,}")
    status = "CLEAN" if c["clean_bill"] else "ACTION REQUIRED"
    print(f"\nOverall status: {status}")

    if finding_rows:
        print(f"\n{'─' * 60}")
        print("FINDINGS WITH REMEDIATION:")
        print(f"{'─' * 60}")
        for row in finding_rows:
            print(f"\n  [{row['severity']}] {row['check_type']} — {row['vendor']}")
            print(f"    Finding: {row['finding']}")
            print(f"    Action:  {row['remediation']}")

    return {
        "summary": summary,
        "match_rows": len(match_rows),
        "finding_rows": len(finding_rows),
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    base = Path(__file__).parent.parent
    parsed = base / "data" / "parsed"
    output = base / "data" / "results"

    run(str(parsed), str(output))
