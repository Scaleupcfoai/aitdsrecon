"""
Parser Agent — TDS Reconciliation MVP
======================================
Parses raw XLSX files from Tally and Form 26 into normalized JSON.

Handles the 2D nature of Tally registers where:
- Row = transaction (date, party, voucher)
- Column = account head (expense category, person, GST component)
- A single value at (row, column) = amount posted to that account

Outputs:
  - parsed_form26.json   — TDS deduction register entries
  - parsed_tally.json     — Tally entries from all 3 registers
"""

import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def to_serializable(obj):
    """Convert datetime objects for JSON serialization."""
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    raise TypeError(f"Type {type(obj)} not serializable")


def clean_name(raw_name: str) -> dict:
    """Parse Form 26 name field: 'Adi Debnath (34); PAN: AAAAA0001A' → {name, id, pan}"""
    if not raw_name:
        return {"name": "", "id": "", "pan": ""}

    match = re.match(r"^(.+?)\s*\((\d+)\);\s*PAN:\s*(\S+)", raw_name)
    if match:
        return {
            "name": match.group(1).strip(),
            "id": match.group(2),
            "pan": match.group(3),
        }
    return {"name": raw_name.strip(), "id": "", "pan": ""}


# ---------------------------------------------------------------------------
# Column classification for dynamic detection
# Instead of hardcoding column names, we detect them from Excel headers
# and classify them using keyword patterns.
# ---------------------------------------------------------------------------

# GST column patterns — match these exactly or by pattern
GST_COLUMN_PATTERNS = {"input c gst", "input s gst", "input i gst", "c gst", "s gst", "i gst", "igst", "cgst", "sgst"}

# Meta columns that are structural, not expense heads
META_COLUMNS_GST_EXP = {
    "Date", "Particulars", "Voucher No.", "Value", "Addl. Cost",
    "Gross Total", "Rounded (+/-)",
}

JOURNAL_META_COLUMNS = {
    "Date", "Particulars", "Voucher No.", "Value", "Gross Total",
}

# Patterns to identify loan/interest party columns (e.g. "X (Loan)")
LOAN_COLUMN_PATTERN = re.compile(r"^(.+?)\s*\(Loan\)$", re.IGNORECASE)

# Patterns to identify director salary columns — detected dynamically
# by checking if the account name appears alongside "Salary & Bonus" or
# "Director's Salary" postings in the same journal entries.
# Not hardcoded — detected at parse time.


def _is_gst_column(col_name: str) -> bool:
    """Check if a column name is a GST column."""
    return col_name.lower().strip() in GST_COLUMN_PATTERNS or \
           "gst" in col_name.lower() and ("input" in col_name.lower() or
           col_name.strip() in {"C GST", "S GST", "I GST"})


def _is_meta_column_gst(col_name: str) -> bool:
    """Check if a column is a meta/structural column in GST Exp Register."""
    return col_name in META_COLUMNS_GST_EXP


def _is_expense_column(col_name: str) -> bool:
    """Any column that is not meta and not GST is an expense head."""
    return not _is_meta_column_gst(col_name) and not _is_gst_column(col_name)


# ---------------------------------------------------------------------------
# 1. Parse Form 26
# ---------------------------------------------------------------------------

def parse_form26(filepath: str) -> list[dict]:
    """
    Parse Form 26 Deduction Register.
    Returns list of dicts with: name, pan, section, amount_paid, date,
    income_tax, surcharge, cess, tax_rate, tax_deducted, tax_date.
    Skips total/subtotal rows.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Deduction Details"]

    entries = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        raw_name = row[1].value  # B column
        section = row[2].value   # C column

        # Skip empty rows, total rows, grand total
        if not raw_name or not section:
            continue
        if "Total" in str(raw_name) or "Grand" in str(raw_name):
            continue

        parsed_name = clean_name(str(raw_name))
        entry = {
            "source": "form26",
            "vendor_name": parsed_name["name"],
            "vendor_id": parsed_name["id"],
            "pan": parsed_name["pan"],
            "section": str(section).strip(),
            "amount_paid": row[3].value,          # D: Amt Paid/Credited
            "amount_paid_date": row[4].value,      # E: Date
            "income_tax": row[5].value or 0,       # F: IT Rs
            "surcharge": row[6].value or 0,        # G
            "cess": row[7].value or 0,             # H
            "tax_rate_pct": row[8].value,          # I: Tax Rate %
            "tax_deducted": row[9].value,          # J: Tax Deducted Rs
            "tax_deducted_date": row[10].value,    # K: Tax Deducted Date
            "non_deduction_reason": row[11].value, # L
        }
        entries.append(entry)

    wb.close()
    return entries


# ---------------------------------------------------------------------------
# 1b. Parse Form 24 (Salary TDS — Section 192)
# ---------------------------------------------------------------------------

def parse_form24(filepath: str) -> list[dict]:
    """
    Parse Form 24Q Deduction Register (Salary TDS).
    Similar structure to Form 26 but no Section column (all are 192)
    and no PAN column. Employee names are abbreviated.

    Returns list of dicts compatible with Form 26 structure.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Deduction Details"]

    entries = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        raw_name = row[1].value  # B column
        amount = row[2].value    # C column

        # Skip empty rows
        if not raw_name or not amount:
            continue
        name_str = str(raw_name).strip()

        # Skip total/subtotal rows (rows without a date are totals)
        date_val = row[3].value
        if not date_val:
            continue
        if "Total" in name_str or "Grand" in name_str:
            continue

        # Compute effective tax rate
        tax_deducted = row[7].value or 0
        tax_rate = round(tax_deducted / amount * 100, 2) if amount else 0

        entry = {
            "source": "form24",
            "vendor_name": name_str,  # abbreviated name like "AD (D1)"
            "vendor_id": "",
            "pan": "",
            "section": "192",
            "amount_paid": amount,            # C: Salary paid
            "amount_paid_date": date_val,     # D: Date
            "income_tax": row[4].value or 0,  # E: IT Rs
            "surcharge": row[5].value or 0,   # F
            "cess": row[6].value or 0,        # G
            "tax_rate_pct": tax_rate,
            "tax_deducted": tax_deducted,     # H: Tax Deducted Rs
            "tax_deducted_date": row[8].value,  # I: Tax Deducted Date
            "non_deduction_reason": row[9].value,  # J
        }
        entries.append(entry)

    wb.close()
    return entries


# ---------------------------------------------------------------------------
# 2. Parse Tally — Journal Register (the complex 2D one)
# ---------------------------------------------------------------------------

def _get_direction(cell) -> str:
    """Extract Cr/Dr direction from a Tally cell's number format.
    Tally encodes direction in the format string, not the sign.
    Returns 'dr' (debit/expense), 'cr' (credit/reversal), or '' (unknown).
    """
    fmt = getattr(cell, 'number_format', '') or ''
    fmt_lower = fmt.lower()
    if ' dr"' in fmt_lower or '"dr"' in fmt_lower:
        return 'dr'
    elif ' cr"' in fmt_lower or '"cr"' in fmt_lower:
        return 'cr'
    return ''


def parse_journal_register(ws) -> list[dict]:
    """
    Parse Journal Register — 68-column 2D register.

    Each row is a journal entry. The columns beyond Gross Total represent
    individual account heads (ledgers). A non-zero value at (row, col) means
    that amount was debited/credited to that account.

    We classify each row by detecting WHICH account columns have values:
    - Interest Paid + Loan column → interest payment (194A relevant)
    - TDS Payable → TDS deduction entry
    - Freight Charges → freight expense (194C relevant for Inland World)
    - Salary & Bonus + person column → salary entry
    - etc.
    """
    # Build column map: column_letter → header_name
    headers = {}
    for cell in ws[7]:
        if cell.value:
            headers[cell.column_letter] = str(cell.value).strip()

    # Invert: header_name → column_index (0-based)
    header_to_idx = {}
    for cell in ws[7]:
        if cell.value:
            header_to_idx[str(cell.value).strip()] = cell.column - 1

    entries = []

    for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
        date_val = row[0].value       # A
        particulars = row[1].value    # B
        voucher_no = row[2].value     # C
        value = row[3].value          # D
        gross_total = row[4].value    # E
        gross_total_dir = _get_direction(row[4])  # Cr/Dr from cell format

        # Skip empty/summary rows
        if particulars and "Grand Total" in str(particulars):
            continue
        if gross_total is None and date_val is None:
            continue

        # Collect all non-zero account columns for this row
        account_postings = {}
        posting_directions = {}  # col_name → 'dr'/'cr'
        for cell in row:
            col_letter = cell.column_letter
            if col_letter in headers and cell.value is not None and cell.value != 0:
                col_name = headers[col_letter]
                if col_name in JOURNAL_META_COLUMNS:
                    continue
                account_postings[col_name] = cell.value
                posting_directions[col_name] = _get_direction(cell)

        if not account_postings:
            continue

        # Classify the entry based on which accounts are hit
        entry_type = _classify_journal_entry(account_postings)

        # For interest entries, extract the loan party dynamically
        # Loan columns match pattern "Name (Loan)" or similar
        loan_party = None
        if "Interest Paid" in account_postings:
            for col_name in account_postings:
                m = LOAN_COLUMN_PATTERN.match(col_name)
                if m:
                    loan_party = m.group(1).strip()
                    break

        entry = {
            "source": "tally_journal",
            "date": date_val,
            "particulars": str(particulars or "").strip(),
            "voucher_no": str(voucher_no or "").strip(),
            "gross_total": gross_total,
            "gross_total_direction": gross_total_dir,  # 'dr'=expense, 'cr'=reversal
            "entry_type": entry_type,
            "account_postings": account_postings,
            "posting_directions": posting_directions,
            "loan_party": loan_party,
        }
        entries.append(entry)

    return entries


def _classify_journal_entry(postings: dict) -> str:
    """Classify a journal entry by which account heads are posted."""
    keys = set(postings.keys())

    if "Interest Paid" in keys:
        return "interest_payment"
    if "TDS Payable" in keys and len(keys) == 1:
        return "tds_deduction"
    if "Freight Charges" in keys:
        return "freight_expense"
    if "Packing Charges" in keys:
        return "packing_expense"
    if "Salary & Bonus" in keys or "Director's Salary" in keys:
        return "salary"
    if "Brokerage and Commission" in keys:
        return "brokerage"
    if "Shop Rent" in keys:
        return "rent"
    if "Consultancy Charges" in keys:
        return "consultancy"
    if "Professonal Charges" in keys:
        return "professional_fees"
    if "Audit Fees" in keys or "Outstanding Audit Fees" in keys:
        return "audit_fees"
    if "TDS Payable" in keys:
        return "tds_deduction"
    # Detect director salary: if any posting key looks like a person name
    # and the entry has no other expense classification, treat as salary.
    # Person name heuristic: 2-4 words, all title case, no special chars
    if any(re.match(r"^[A-Z][a-z]+(?: [A-Z][a-z]+){1,3}$", k) for k in keys):
        return "salary"
    if "Cash Discount" in keys or "Discount (CD)" in keys:
        return "discount"

    return "other"


# ---------------------------------------------------------------------------
# 3. Parse Tally — Purchase GST Exp Register (2D expense register)
# ---------------------------------------------------------------------------

def parse_purchase_gst_exp_register(ws) -> list[dict]:
    """
    Parse Purchase GST Expense Register — 42-column 2D register.

    Each row is an expense voucher. The columns represent expense categories.
    A non-zero value at (row, col) means that amount was booked under that
    expense category. GST columns (CGST, SGST, IGST) are separate.

    Key for 194C matching: vendors like Amrita Icons, Anderson Technology
    have expenses here (Packing, Advertisement etc.) with GST breakup.
    """
    headers = {}
    for cell in ws[7]:
        if cell.value:
            headers[cell.column_letter] = str(cell.value).strip()

    entries = []

    for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
        date_val = row[0].value
        particulars = row[1].value
        voucher_no = row[2].value

        if particulars and "Grand Total" in str(particulars):
            continue
        if date_val is None and particulars is None:
            continue

        # Collect values by column type
        gross_total = None
        gross_total_dir = ''
        expense_heads = {}
        gst_amounts = {}
        rounding = 0
        value = None
        addl_cost = None

        for cell in row:
            col_letter = cell.column_letter
            if col_letter not in headers or cell.value is None or cell.value == 0:
                continue

            col_name = headers[col_letter]

            if col_name == "Gross Total":
                gross_total = cell.value
                gross_total_dir = _get_direction(cell)
            elif col_name == "Value":
                value = cell.value
            elif col_name == "Addl. Cost":
                addl_cost = cell.value
            elif col_name == "Rounded (+/-)":
                rounding = cell.value
            elif _is_gst_column(col_name):
                gst_amounts[col_name] = cell.value
            elif _is_expense_column(col_name):
                expense_heads[col_name] = cell.value

        if not expense_heads and not gst_amounts:
            continue

        # Compute base amount (sum of expense heads, before GST)
        base_amount = sum(expense_heads.values())
        total_gst = sum(gst_amounts.values())

        entry = {
            "source": "tally_purchase_gst_exp",
            "date": date_val,
            "particulars": str(particulars or "").strip(),
            "voucher_no": str(voucher_no or "").strip(),
            "value": value,
            "gross_total": gross_total,
            "gross_total_direction": gross_total_dir,
            "base_amount": base_amount,
            "total_gst": total_gst,
            "rounding": rounding,
            "expense_heads": expense_heads,
            "gst_breakup": gst_amounts,
        }
        entries.append(entry)

    return entries


# ---------------------------------------------------------------------------
# 4. Parse Tally — Purchase Register (goods purchases)
# ---------------------------------------------------------------------------

def parse_purchase_register(ws) -> list[dict]:
    """
    Parse Purchase Register — goods purchases with GST breakup.
    Simpler structure: each row is a purchase with value, GST columns,
    discount, freight, etc.
    """
    headers = {}
    for cell in ws[7]:
        if cell.value:
            headers[cell.column_letter] = str(cell.value).strip()

    entries = []

    for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
        date_val = row[0].value
        particulars = row[1].value
        voucher_no = row[2].value

        if particulars and "Grand Total" in str(particulars):
            continue
        if date_val is None and particulars is None:
            continue

        # Collect all non-meta column values
        amounts = {}
        for cell in row:
            col_letter = cell.column_letter
            if col_letter not in headers or cell.value is None:
                continue
            col_name = headers[col_letter]
            if col_name not in ("Date", "Particulars", "Voucher No."):
                amounts[col_name] = cell.value

        if not amounts:
            continue

        # Extract key fields
        purchase_value = amounts.get("Value") or amounts.get("Purchase Account", 0)
        cgst = amounts.get("Input C GST", 0)
        sgst = amounts.get("Input S GST", 0)
        igst = amounts.get("Input I GST", 0)
        gross_total = amounts.get("Gross Total", 0)
        discount = amounts.get("Discount Received From Purchase", 0)

        entry = {
            "source": "tally_purchase",
            "date": date_val,
            "particulars": str(particulars or "").strip(),
            "voucher_no": str(voucher_no or "").strip(),
            "purchase_value": purchase_value,
            "gross_total": gross_total,
            "discount": discount,
            "cgst": cgst,
            "sgst": sgst,
            "igst": igst,
            "total_gst": (cgst or 0) + (sgst or 0) + (igst or 0),
            "all_amounts": amounts,
        }
        entries.append(entry)

    return entries


# ---------------------------------------------------------------------------
# Main — Run Parser Agent
# ---------------------------------------------------------------------------

def run(form26_path: str, tally_path: str, output_dir: str, form24_path: str | None = None):
    """Run the parser agent end-to-end."""

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # --- Parse Form 26 ---
    print("[Parser] Parsing Form 26...")
    form26_entries = parse_form26(form26_path)
    print(f"  → {len(form26_entries)} entries extracted")

    # --- Parse Form 24 (if provided) ---
    form24_entries = []
    if form24_path and Path(form24_path).exists():
        print("\n[Parser] Parsing Form 24 (Salary TDS)...")
        form24_entries = parse_form24(form24_path)
        print(f"  → {len(form24_entries)} salary TDS entries extracted")
        # Merge into form26_entries — they share the same structure
        form26_entries.extend(form24_entries)
        print(f"  → Combined: {len(form26_entries)} total entries (Form 26 + Form 24)")

    # Count by section
    sections = {}
    for e in form26_entries:
        s = e["section"]
        sections[s] = sections.get(s, 0) + 1
    for s, count in sorted(sections.items()):
        print(f"    {s}: {count} entries")

    # --- Parse Tally ---
    print("\n[Parser] Parsing Tally extract...")
    # Open WITHOUT data_only so we can read Cr/Dr number formats from cells
    wb = openpyxl.load_workbook(tally_path)

    print("  Parsing Journal Register...")
    journal_entries = parse_journal_register(wb["Journal Register"])
    print(f"  → {len(journal_entries)} entries")
    # Count by type
    types = {}
    for e in journal_entries:
        t = e["entry_type"]
        types[t] = types.get(t, 0) + 1
    for t, count in sorted(types.items()):
        print(f"    {t}: {count}")

    print("  Parsing Purchase GST Exp Register...")
    gst_exp_entries = parse_purchase_gst_exp_register(wb["Purchase GST Exp. Register"])
    print(f"  → {len(gst_exp_entries)} entries")

    print("  Parsing Purchase Register...")
    purchase_entries = parse_purchase_register(wb["Purchase Register"])
    print(f"  → {len(purchase_entries)} entries")

    wb.close()

    # --- Build output ---
    form26_output = {
        "source_file": str(form26_path),
        "parsed_at": datetime.now().isoformat(),
        "total_entries": len(form26_entries),
        "entries_by_section": sections,
        "entries": form26_entries,
    }

    tally_output = {
        "source_file": str(tally_path),
        "parsed_at": datetime.now().isoformat(),
        "journal_register": {
            "total_entries": len(journal_entries),
            "entries_by_type": types,
            "entries": journal_entries,
        },
        "purchase_gst_exp_register": {
            "total_entries": len(gst_exp_entries),
            "entries": gst_exp_entries,
        },
        "purchase_register": {
            "total_entries": len(purchase_entries),
            "entries": purchase_entries,
        },
    }

    # --- Write output ---
    form26_out = output_dir / "parsed_form26.json"
    tally_out = output_dir / "parsed_tally.json"

    with open(form26_out, "w") as f:
        json.dump(form26_output, f, indent=2, default=to_serializable)
    print(f"\n[Parser] Wrote {form26_out}")

    with open(tally_out, "w") as f:
        json.dump(tally_output, f, indent=2, default=to_serializable)
    print(f"[Parser] Wrote {tally_out}")

    # --- Summary for downstream agents ---
    print("\n" + "=" * 60)
    print("PARSER AGENT — SUMMARY")
    print("=" * 60)

    print(f"\nForm 26: {len(form26_entries)} deduction entries")
    print(f"  194A (Interest): {sections.get('194A', 0)} entries")
    print(f"  194C (Contractor): {sections.get('194C', 0)} entries")
    print(f"  194H (Commission): {sections.get('194H', 0)} entries")
    print(f"  194J(b) (Professional): {sections.get('194J(b)', 0)} entries")
    print(f"  194Q (Purchase): {sections.get('194Q', 0)} entries")

    print(f"\nTally Journal Register: {len(journal_entries)} entries")
    print(f"  Interest payments: {types.get('interest_payment', 0)}")
    print(f"  TDS deductions: {types.get('tds_deduction', 0)}")
    print(f"  Freight expenses: {types.get('freight_expense', 0)}")

    print(f"\nTally Purchase GST Exp: {len(gst_exp_entries)} entries")
    print(f"Tally Purchase Register: {len(purchase_entries)} entries")

    return form26_output, tally_output


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Default paths for HPC data
    base = Path(__file__).parent.parent.parent
    form26 = base / "data" / "hpc" / "Form 26 - Deduction Register....xlsx"
    tally = base / "data" / "hpc" / "Tally extract.xlsx"
    output = base / "tds-recon" / "data" / "parsed"

    if len(sys.argv) >= 3:
        form26 = sys.argv[1]
        tally = sys.argv[2]
    if len(sys.argv) >= 4:
        output = sys.argv[3]

    run(str(form26), str(tally), str(output))
