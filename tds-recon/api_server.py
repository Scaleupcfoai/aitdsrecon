"""
FastAPI bridge for TDS Reconciliation agents.
Wraps the Python pipeline and exposes it to the React UI.

Run: uvicorn api_server:app --reload --port 8000
"""

import json
import os
import queue
import sys
import threading
import time
from pathlib import Path

import httpx
import shutil
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")  # Reads .env from same folder as api_server.py

from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel

# Add parent to path so agents module is importable
sys.path.insert(0, str(Path(__file__).parent))

app = FastAPI(title="TDS Recon API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE = Path(__file__).parent
RESULTS_DIR = BASE / "data" / "results"
RULES_DIR = BASE / "data" / "rules"
PARSED_DIR = BASE / "data" / "parsed"
UPLOADS_DIR = BASE / "data" / "uploads"


@app.get("/api/status")
def get_status():
    """Check if parsed data and results exist."""
    return {
        "parsed_ready": (PARSED_DIR / "parsed_form26.json").exists(),
        "results_ready": (RESULTS_DIR / "match_results.json").exists(),
        "rules_ready": (RULES_DIR / "learned_rules.json").exists(),
    }


@app.post("/api/upload")
async def upload_files(
    form26: UploadFile = File(...),
    tally: UploadFile = File(...),
    form24: UploadFile | None = File(None),
):
    """Upload Form 26, Tally, and optionally Form 24 XLSX files."""
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

    form26_path = UPLOADS_DIR / "form26.xlsx"
    tally_path = UPLOADS_DIR / "tally.xlsx"

    with open(form26_path, "wb") as f:
        shutil.copyfileobj(form26.file, f)
    with open(tally_path, "wb") as f:
        shutil.copyfileobj(tally.file, f)

    result = {
        "status": "uploaded",
        "form26": form26.filename,
        "tally": tally.filename,
    }

    if form24 and form24.filename:
        form24_path = UPLOADS_DIR / "form24.xlsx"
        with open(form24_path, "wb") as f:
            shutil.copyfileobj(form24.file, f)
        result["form24"] = form24.filename

    return result


@app.post("/api/preview-columns")
async def preview_columns(
    form26: UploadFile = File(...),
    tally: UploadFile = File(...),
    form24: UploadFile | None = File(None),
):
    """Lightweight preview: read headers, sample rows, and entry counts from uploaded files.

    Returns enough info for the UI to show a confirmation card before running the pipeline.
    Files are saved to disk (same as /api/upload) so the pipeline can use them directly.
    """
    import openpyxl

    # Save files to disk (reuse upload logic)
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    form26_path = UPLOADS_DIR / "form26.xlsx"
    tally_path = UPLOADS_DIR / "tally.xlsx"

    with open(form26_path, "wb") as f:
        shutil.copyfileobj(form26.file, f)
    with open(tally_path, "wb") as f:
        shutil.copyfileobj(tally.file, f)

    form24_saved = False
    if form24 and form24.filename:
        form24_path = UPLOADS_DIR / "form24.xlsx"
        with open(form24_path, "wb") as f:
            shutil.copyfileobj(form24.file, f)
        form24_saved = True

    preview = {"form26": {}, "tally": {}}

    # --- Form 26 preview ---
    try:
        wb = openpyxl.load_workbook(str(form26_path), data_only=True)
        sheet_names = wb.sheetnames
        ws = wb["Deduction Details"]

        # Read header row (row 4 has column labels)
        headers = []
        for cell in ws[4]:
            if cell.value:
                headers.append(str(cell.value).strip())

        # Count entries and sections (lightweight scan)
        sections = {}
        sample_rows = []
        total = 0
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            raw_name = row[1].value
            section = row[2].value
            if not raw_name or not section:
                continue
            if "Total" in str(raw_name) or "Grand" in str(raw_name):
                continue
            total += 1
            s = str(section).strip()
            sections[s] = sections.get(s, 0) + 1
            if len(sample_rows) < 3:
                sample_rows.append({
                    "vendor": str(raw_name).split("(")[0].strip()[:40],
                    "section": s,
                    "amount": row[3].value,
                    "tax_deducted": row[9].value,
                })
        wb.close()

        preview["form26"] = {
            "filename": form26.filename,
            "sheet_names": sheet_names,
            "detected_sheet": "Deduction Details",
            "headers": headers,
            "total_entries": total,
            "sections": sections,
            "sample_rows": sample_rows,
        }
    except Exception as e:
        preview["form26"] = {"filename": form26.filename, "error": str(e)}

    # --- Tally preview ---
    try:
        wb = openpyxl.load_workbook(str(tally_path), data_only=True)
        sheet_names = wb.sheetnames
        registers = {}

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            # Read headers from row 7 (Tally standard)
            headers = []
            for cell in ws[7]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            # Count data rows
            row_count = 0
            for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
                if row[1].value and "Grand Total" not in str(row[1].value):
                    row_count += 1
            registers[sheet_name] = {
                "headers": headers,
                "entry_count": row_count,
            }
        wb.close()

        preview["tally"] = {
            "filename": tally.filename,
            "sheet_names": sheet_names,
            "registers": registers,
        }
    except Exception as e:
        preview["tally"] = {"filename": tally.filename, "error": str(e)}

    # --- Form 24 preview (if uploaded) ---
    if form24_saved:
        try:
            form24_path = UPLOADS_DIR / "form24.xlsx"
            wb = openpyxl.load_workbook(str(form24_path), data_only=True)
            ws = wb["Deduction Details"]
            emp_count = 0
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
                if row[1].value and row[3].value and "Total" not in str(row[1].value):
                    emp_count += 1
            wb.close()
            preview["form24"] = {
                "filename": form24.filename,
                "total_entries": emp_count,
                "section": "192 (Salary)",
            }
        except Exception as e:
            preview["form24"] = {"filename": form24.filename, "error": str(e)}

    return preview


@app.get("/api/preview-columns/existing")
def preview_existing_columns():
    """Preview columns from already-parsed data on disk (for 'Use Existing Data' flow)."""
    preview = {"form26": {}, "tally": {}}

    # Read parsed Form 26
    f26_path = PARSED_DIR / "parsed_form26.json"
    if f26_path.exists():
        with open(f26_path) as f:
            f26 = json.load(f)
        entries = f26.get("entries", [])
        sections = f26.get("entries_by_section", {})
        sample_rows = []
        for e in entries[:3]:
            sample_rows.append({
                "vendor": (e.get("vendor_name") or "Unknown")[:40],
                "section": e.get("section", ""),
                "amount": e.get("amount_paid", 0),
                "tax_deducted": e.get("tax_deducted", 0),
            })
        preview["form26"] = {
            "filename": Path(f26.get("source_file", "Form 26")).name,
            "sheet_names": ["Deduction Details"],
            "detected_sheet": "Deduction Details",
            "headers": ["Vendor Name", "Section", "Amt Paid", "Date", "IT Rs", "Surcharge", "Cess", "Tax Rate %", "Tax Deducted", "Tax Date"],
            "total_entries": f26.get("total_entries", len(entries)),
            "sections": sections,
            "sample_rows": sample_rows,
        }
    else:
        preview["form26"] = {"error": "No parsed Form 26 data found. Upload files first."}

    # Read parsed Tally
    tally_path = PARSED_DIR / "parsed_tally.json"
    if tally_path.exists():
        with open(tally_path) as f:
            tally = json.load(f)
        registers = {}
        for reg_name in ["journal_register", "purchase_gst_exp_register", "purchase_register"]:
            reg = tally.get(reg_name, {})
            if reg:
                display_name = reg_name.replace("_", " ").title().replace("Gst Exp", "GST Exp.")
                registers[display_name] = {
                    "entry_count": reg.get("total_entries", len(reg.get("entries", []))),
                    "headers": list(reg.get("entries_by_type", {}).keys()) if "entries_by_type" in reg else [],
                }
        preview["tally"] = {
            "filename": Path(tally.get("source_file", "Tally Extract")).name,
            "sheet_names": list(registers.keys()),
            "registers": registers,
        }
    else:
        preview["tally"] = {"error": "No parsed Tally data found. Upload files first."}

    return preview


@app.get("/api/run/stream/upload")
def run_pipeline_with_upload():
    """Run the full pipeline on uploaded files with real-time SSE.

    Parses the uploaded XLSX files first, then runs matcher → checker → reporter.
    """
    form26_path = UPLOADS_DIR / "form26.xlsx"
    tally_path = UPLOADS_DIR / "tally.xlsx"

    if not form26_path.exists() or not tally_path.exists():
        def error_gen():
            yield f"data: {json.dumps({'type': 'error', 'agent': 'Upload', 'message': 'Upload Form 26 and Tally files first'})}\n\n"
        return StreamingResponse(error_gen(), media_type="text/event-stream")

    event_queue = queue.Queue()

    def on_event(event):
        event_queue.put(event)

    def run_in_thread():
        from reconcile import run_pipeline as _run

        form24_path = UPLOADS_DIR / "form24.xlsx"
        f24 = str(form24_path) if form24_path.exists() else None
        # Pass callback directly — reconcile.py resets the logger internally
        result = _run(str(form26_path), str(tally_path), form24_path=f24, event_callback=on_event)

        results_data = _load_results()
        event_queue.put({
            "type": "pipeline_complete",
            "agent": "Pipeline",
            "message": f"Complete in {result.get('elapsed_s', 0)}s",
            "results": results_data,
        })
        event_queue.put(None)

    thread = threading.Thread(target=run_in_thread, daemon=True)
    thread.start()

    return StreamingResponse(
        _sse_generator(event_queue),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/run/stream")
def run_pipeline_stream():
    """Run the full pipeline with real-time SSE streaming.

    Each event is sent as it happens — no batching, no fake delays.
    Final event includes the full results payload.
    """
    event_queue = queue.Queue()

    def on_event(event):
        event_queue.put(event)

    def run_in_thread():
        from reconcile import run_pipeline as _run

        # Pass callback directly — reconcile.py resets the logger internally
        result = _run(event_callback=on_event)

        # Send final results as a special event
        results_data = _load_results()
        event_queue.put({
            "type": "pipeline_complete",
            "agent": "Pipeline",
            "message": f"Complete in {result.get('elapsed_s', 0)}s",
            "results": results_data,
        })
        event_queue.put(None)  # Sentinel to end stream

    thread = threading.Thread(target=run_in_thread, daemon=True)
    thread.start()

    return StreamingResponse(
        _sse_generator(event_queue),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/run")
def run_pipeline():
    """Execute the full pipeline (non-streaming fallback)."""
    from reconcile import run_pipeline as _run
    result = _run()
    result["results"] = _load_results()
    return result


# ---------------------------------------------------------------------------
# Individual Agent Endpoints
# ---------------------------------------------------------------------------

@app.post("/api/run/parser")
def run_parser():
    """Run only the Parser Agent."""
    from agents.event_logger import reset_logger
    logger = reset_logger()
    start = time.time()
    logger.agent_start("Parser Agent", "Starting Parser Agent...")

    if not (PARSED_DIR / "parsed_form26.json").exists():
        logger.error("Parser Agent", "parsed_form26.json not found")
        return {"events": logger.get_events(), "error": "Missing parsed data"}
    if not (PARSED_DIR / "parsed_tally.json").exists():
        logger.error("Parser Agent", "parsed_tally.json not found")
        return {"events": logger.get_events(), "error": "Missing parsed data"}

    with open(PARSED_DIR / "parsed_form26.json") as f:
        f26 = json.load(f)
    with open(PARSED_DIR / "parsed_tally.json") as f:
        tally = json.load(f)

    f26_count = len(f26.get("entries", []))
    sections = set(e["section"] for e in f26.get("entries", []))
    jr_count = len(tally.get("journal_register", {}).get("entries", []))
    gst_count = len(tally.get("purchase_gst_exp_register", {}).get("entries", []))
    pr_count = len(tally.get("purchase_register", {}).get("entries", []))

    logger.detail("Parser Agent", f"Form 26: {f26_count} entries across {len(sections)} sections")
    logger.detail("Parser Agent", f"Sections found: {', '.join(sorted(sections))}")
    logger.detail("Parser Agent", f"Tally Journal Register: {jr_count} entries")
    logger.detail("Parser Agent", f"Tally GST Expense Register: {gst_count} entries")
    logger.detail("Parser Agent", f"Tally Purchase Register: {pr_count} entries")

    elapsed = time.time() - start
    logger.agent_done("Parser Agent", f"Parsing complete ({elapsed:.1f}s)")
    return {"events": logger.get_events(), "elapsed_s": round(elapsed, 2)}


@app.post("/api/run/matcher")
def run_matcher():
    """Run only the Matcher Agent."""
    from agents.event_logger import reset_logger
    from agents.matcher_agent import run as matcher_run
    logger = reset_logger()
    start = time.time()

    logger.agent_start("Matcher Agent", "Starting Matcher Agent...")
    match_results = matcher_run(str(PARSED_DIR), str(RESULTS_DIR), rules_dir=str(RULES_DIR))

    summary = match_results.get("summary", {})
    matched = summary.get("form26_matched", 0)
    total = summary.get("form26_total", 0)
    pct = (matched / total * 100) if total > 0 else 0
    logger.success("Matcher Agent", f"Result: {matched}/{total} matched ({pct:.0f}%)")

    elapsed = time.time() - start
    logger.agent_done("Matcher Agent", f"Matching complete ({elapsed:.1f}s)")
    return {"events": logger.get_events(), "elapsed_s": round(elapsed, 2), "results": {"match_results": match_results}}


@app.post("/api/run/checker")
def run_checker():
    """Run only the TDS Checker Agent."""
    from agents.event_logger import reset_logger
    from agents.tds_checker_agent import run as checker_run
    logger = reset_logger()
    start = time.time()

    if not (RESULTS_DIR / "match_results.json").exists():
        return {"events": [], "error": "Run Matcher first"}

    logger.agent_start("TDS Checker", "Starting TDS Checker Agent...")
    checker_results = checker_run(str(PARSED_DIR), str(RESULTS_DIR))
    findings = checker_results.get("findings", [])
    errors = [f for f in findings if f.get("severity") == "error"]
    warnings = [f for f in findings if f.get("severity") == "warning"]
    logger.success("TDS Checker", f"Complete: {len(errors)} errors, {len(warnings)} warnings")

    elapsed = time.time() - start
    logger.agent_done("TDS Checker", f"Compliance checks complete ({elapsed:.1f}s)")
    return {"events": logger.get_events(), "elapsed_s": round(elapsed, 2), "results": {"checker_results": checker_results}}


@app.post("/api/run/reporter")
def run_reporter():
    """Run only the Reporter Agent."""
    from agents.event_logger import reset_logger
    from agents.reporter_agent import run as reporter_run
    logger = reset_logger()
    start = time.time()

    if not (RESULTS_DIR / "match_results.json").exists():
        return {"events": [], "error": "Run Matcher first"}

    logger.agent_start("Reporter Agent", "Generating reports...")
    report = reporter_run(str(PARSED_DIR), str(RESULTS_DIR))

    elapsed = time.time() - start
    logger.agent_done("Reporter Agent", f"Reports generated ({elapsed:.1f}s)")
    return {"events": logger.get_events(), "elapsed_s": round(elapsed, 2), "results": {"reconciliation_summary": report.get("summary", {})}}


# ---------------------------------------------------------------------------
# Download Reports
# ---------------------------------------------------------------------------

@app.get("/api/download/{filename}")
def download_report(filename: str):
    """Download a report file (CSV or JSON)."""
    allowed = {
        "tds_recon_report.xlsx",
        "reconciliation_report.csv",
        "findings_report.csv",
        "reconciliation_summary.json",
        "match_results.json",
        "checker_results.json",
    }
    if filename not in allowed:
        return {"error": f"File not available: {filename}"}
    fpath = RESULTS_DIR / filename
    if not fpath.exists():
        return {"error": f"File not found. Run the pipeline first."}
    return FileResponse(
        path=str(fpath),
        filename=filename,
        media_type="application/octet-stream",
    )


# ---------------------------------------------------------------------------
# Results + Review
# ---------------------------------------------------------------------------

@app.get("/api/results")
def get_results():
    return _load_results()


@app.get("/api/rules")
def get_rules():
    from agents.learning_agent import load_rules, summarize_rules
    db = load_rules(str(RULES_DIR))
    summary = summarize_rules(str(RULES_DIR))
    return {"rules": db, "summary": summary}


class AnswerRequest(BaseModel):
    question_id: str
    selected: list[str] = []
    text_input: str | None = None


@app.post("/api/answer")
def submit_answer(request: AnswerRequest):
    """Submit an answer to a pipeline decision point question."""
    from agents.event_logger import get_logger
    logger = get_logger()
    logger.set_answer(request.question_id, {
        "selected": request.selected,
        "text_input": request.text_input,
    })
    return {"status": "ok", "question_id": request.question_id}


class ReviewDecision(BaseModel):
    vendor: str
    decision: str
    params: dict = {}
    reason: str = ""


class ReviewRequest(BaseModel):
    decisions: list[ReviewDecision]


@app.post("/api/review")
def submit_review(request: ReviewRequest):
    from agents.learning_agent import apply_corrections
    decisions = [d.model_dump() for d in request.decisions]
    result = apply_corrections(str(RULES_DIR), str(RESULTS_DIR), decisions)
    return result


# ---------------------------------------------------------------------------
# LLM Chat
# ---------------------------------------------------------------------------

GEMINI_API_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent"

SYSTEM_PROMPT = """You are Lekha AI, a TDS reconciliation assistant for Indian tax compliance.
You are helping a CA/accountant review TDS reconciliation results for FY 2024-25 (AY 2025-26).

Rules:
- Answer using ONLY the reconciliation data provided below. Do not make up numbers.
- Use ₹ with Indian number format (e.g. ₹1,24,533).
- Be specific: name vendors, sections, amounts, rates when answering.
- If the data doesn't contain what was asked, say so clearly.
- Keep responses concise (3-5 sentences unless the user asks for detail).
- You can NOT run the pipeline, modify data, or trigger actions. You only answer questions about the results.

Here is the complete reconciliation data:

"""


def _build_compact_results() -> str | None:
    """Build a compact JSON string of pipeline results for the LLM context."""
    summary_path = RESULTS_DIR / "reconciliation_summary.json"
    checker_path = RESULTS_DIR / "checker_results.json"
    match_path = RESULTS_DIR / "match_results.json"

    if not summary_path.exists():
        return None

    with open(summary_path) as f:
        summary = json.load(f)

    checker = {}
    if checker_path.exists():
        with open(checker_path) as f:
            checker = json.load(f)

    compact_matches = []
    if match_path.exists():
        with open(match_path) as f:
            raw = json.load(f)
        for m in raw.get("matches", []):
            f26 = m.get("form26_entry", {})
            compact_matches.append({
                "vendor": f26.get("vendor_name", ""),
                "pan": f26.get("pan", ""),
                "section": f26.get("section", ""),
                "amount_paid": f26.get("amount_paid", 0),
                "tax_deducted": f26.get("tax_deducted", 0),
                "tax_rate_pct": f26.get("tax_rate_pct", 0),
                "confidence": m.get("confidence", 0),
                "match_type": m.get("pass_name", ""),
            })

    context = {
        "reconciliation_summary": summary,
        "compliance_findings": checker.get("findings", []),
        "compliance_summary": checker.get("summary", {}),
        "matched_entries": compact_matches,
    }
    return json.dumps(context, default=str)


class ChatRequest(BaseModel):
    message: str
    history: list[dict] = []


@app.post("/api/chat")
async def chat(request: ChatRequest):
    """LLM chat endpoint. Sends user message + pipeline results to Gemini 2.5 Flash."""
    api_key = os.environ.get("GEMINI_API_KEY", "")
    if not api_key:
        return {"error": "GEMINI_API_KEY not set", "response": "I'm not connected to my AI backend right now. Please set the GEMINI_API_KEY environment variable and restart the server."}

    # Build context from pipeline results
    compact_results = _build_compact_results()
    if not compact_results:
        return {"response": "No reconciliation data available yet. Please run the pipeline first, then ask me questions about the results."}

    system_prompt = SYSTEM_PROMPT + compact_results

    # Build conversation: history + new message
    contents = []
    for msg in request.history[-10:]:  # Last 10 messages
        role = "user" if msg.get("role") == "user" else "model"
        text = msg.get("content", "")
        if text:
            contents.append({"role": role, "parts": [{"text": text}]})
    contents.append({"role": "user", "parts": [{"text": request.message}]})

    payload = {
        "contents": contents,
        "systemInstruction": {"parts": [{"text": system_prompt}]},
        "generationConfig": {
            "temperature": 0.3,
            "maxOutputTokens": 1024,
        },
    }

    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(
                f"{GEMINI_API_URL}?key={api_key}",
                json=payload,
                headers={"Content-Type": "application/json"},
            )
        if resp.status_code != 200:
            error_detail = resp.text[:200]
            return {"error": f"Gemini API error ({resp.status_code})", "response": f"I encountered an issue connecting to my AI backend. Error: {error_detail}"}

        data = resp.json()
        candidates = data.get("candidates", [])
        if not candidates:
            return {"response": "I couldn't generate a response. Please try rephrasing your question."}

        text = candidates[0].get("content", {}).get("parts", [{}])[0].get("text", "")
        return {"response": text}

    except httpx.TimeoutException:
        return {"response": "The request timed out. Please try again with a simpler question."}
    except Exception as e:
        return {"error": str(e), "response": "Something went wrong while processing your question. Please try again."}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _sse_generator(event_queue):
    """Yield SSE events from a queue until sentinel (None) is received."""
    while True:
        try:
            event = event_queue.get(timeout=30)
            if event is None:
                break
            yield f"data: {json.dumps(event, default=str)}\n\n"
        except queue.Empty:
            yield f"data: {json.dumps({'type': 'keepalive'})}\n\n"


def _load_results() -> dict:
    results = {}
    for fname in ["match_results.json", "checker_results.json", "reconciliation_summary.json"]:
        fpath = RESULTS_DIR / fname
        if fpath.exists():
            with open(fpath) as f:
                results[fname.replace(".json", "")] = json.load(f)
    return results


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
