"""
Intelligent Column Mapper — takes any XLSX/CSV and maps columns to fields.

Approach 1 (no cross-verification):
1. Fuzzy matching → confidence score per column
2. confidence >= 0.8 → auto-map (done, no LLM needed)
3. confidence < 0.8 → send ONLY uncertain columns to LLM
4. LLM responds → done
5. If LLM underconfident → flag for human review

Saves confirmed mappings to column_map table for reuse.

Usage:
    from app.services.column_mapper import ColumnMapper
    mapper = ColumnMapper(repo, llm)
    result = mapper.map_file("path/to/file.xlsx", company_id="abc-123")
"""

import csv
import re
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

from app.services.llm_client import LLMClient
from app.services.llm_prompts import PARSER_COLUMN_MAP_SYSTEM, PARSER_COLUMN_MAP_PROMPT


# ═══════════════════════════════════════════════════════════
# Known field patterns
# ═══════════════════════════════════════════════════════════

TDS_FIELDS = {
    "party_name": {
        "keywords": ["name", "party", "vendor", "deductee", "payee", "particulars"],
        "description": "Name of the vendor/party",
    },
    "pan": {
        "keywords": ["pan", "pan no", "pan number", "permanent account"],
        "description": "PAN number (10-char alphanumeric)",
    },
    "tds_section": {
        "keywords": ["section", "tds section", "sec"],
        "description": "TDS section (194A, 194C, etc.)",
    },
    "gross_amount": {
        "keywords": ["amount paid", "amt paid", "gross amount", "amount credited",
                     "amt. paid", "amt paid credited", "payment amount"],
        "description": "Gross amount paid or credited",
    },
    "tds_amount": {
        "keywords": ["tax deducted", "tds amount", "tds deducted", "tax amount",
                     "income tax", "tds amt"],
        "description": "TDS amount deducted",
    },
    "date_of_deduction": {
        "keywords": ["date", "deduction date", "tax date", "payment date",
                     "date of deduction", "date of payment"],
        "description": "Date of TDS deduction or payment",
    },
    "tax_rate": {
        "keywords": ["rate", "tax rate", "tds rate", "rate %", "percentage"],
        "description": "TDS rate percentage",
    },
    "certificate_number": {
        "keywords": ["certificate", "cert no", "certificate no", "tds certificate"],
        "description": "TDS certificate number",
    },
}

LEDGER_FIELDS = {
    "party_name": {
        "keywords": ["particulars", "party", "vendor", "name", "ledger"],
        "description": "Vendor/party name",
    },
    "amount": {
        "keywords": ["amount", "gross total", "total", "value", "debit", "credit"],
        "description": "Transaction amount",
    },
    "invoice_number": {
        "keywords": ["voucher", "voucher no", "invoice", "bill no", "ref"],
        "description": "Invoice or voucher number",
    },
    "invoice_date": {
        "keywords": ["date"],
        "description": "Transaction date",
    },
    "expense_type": {
        "keywords": ["expense", "account", "head", "nature", "type", "ledger"],
        "description": "Expense head / account type",
    },
}

AUTO_MAP_THRESHOLD = 0.8  # >= this confidence = auto-map, no LLM needed


# ═══════════════════════════════════════════════════════════
# Step 1: Read file headers + sample data
# ═══════════════════════════════════════════════════════════

def read_file_headers(filepath: str) -> list[dict]:
    """Read headers and sample data from XLSX or CSV."""
    path = Path(filepath)
    if path.suffix.lower() in (".xlsx", ".xls"):
        return _read_xlsx_headers(filepath)
    elif path.suffix.lower() == ".csv":
        return [_read_csv_headers(filepath)]
    else:
        raise ValueError(f"Unsupported file type: {path.suffix}")


def _read_xlsx_headers(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        # Find header row — row with most text cells in first 15 rows
        best_row, best_count = 1, 0
        for row_num in range(1, min(16, ws.max_row + 1)):
            text_count = sum(
                1 for col in range(1, min(ws.max_column + 1, 100))
                if ws.cell(row_num, col).value and isinstance(ws.cell(row_num, col).value, str)
                and not ws.cell(row_num, col).value.replace(".", "").replace(",", "").isdigit()
            )
            if text_count > best_count:
                best_count = text_count
                best_row = row_num

        headers = [
            {"col_index": col, "col_letter": openpyxl.utils.get_column_letter(col),
             "name": str(ws.cell(best_row, col).value).strip()}
            for col in range(1, ws.max_column + 1)
            if ws.cell(best_row, col).value
        ]

        # 3 sample data rows after header
        sample_rows = []
        for row_num in range(best_row + 1, min(best_row + 4, ws.max_row + 1)):
            row_data = [str(ws.cell(row_num, col).value) if ws.cell(row_num, col).value is not None else ""
                        for col in range(1, ws.max_column + 1)]
            if any(row_data) and not any("total" in str(v).lower() for v in row_data[:3]):
                sample_rows.append(row_data)

        if headers:
            results.append({
                "sheet_name": sheet_name, "header_row": best_row,
                "headers": headers, "sample_rows": sample_rows[:3],
                "total_rows": ws.max_row - best_row, "total_cols": len(headers),
            })
    wb.close()
    return results


def _read_csv_headers(filepath: str) -> dict:
    with open(filepath, encoding="utf-8-sig") as f:
        rows = [row for i, row in zip(range(6), csv.reader(f))]

    header_idx = 0
    for i, row in enumerate(rows):
        if sum(1 for v in row if v and not v.replace(".", "").replace(",", "").isdigit()) >= 3:
            header_idx = i
            break

    headers = [{"col_index": j + 1, "col_letter": chr(65 + j) if j < 26 else f"C{j}", "name": v.strip()}
               for j, v in enumerate(rows[header_idx]) if v.strip()]
    return {
        "sheet_name": "CSV", "header_row": header_idx + 1,
        "headers": headers, "sample_rows": rows[header_idx + 1: header_idx + 4],
        "total_rows": "unknown", "total_cols": len(headers),
    }


# ═══════════════════════════════════════════════════════════
# Step 2: Fuzzy matching
# ═══════════════════════════════════════════════════════════

def fuzzy_match_columns(headers: list[dict], field_definitions: dict) -> list[dict]:
    """Score each column against known fields using keyword similarity."""
    results = []
    for header in headers:
        col_name = header["name"].lower().strip()
        best_field, best_score, best_method = None, 0.0, "none"

        for field_name, field_def in field_definitions.items():
            for keyword in field_def["keywords"]:
                kw = keyword.lower()
                if col_name == kw:
                    score, method = 1.0, "exact"
                elif kw in col_name:
                    score, method = 0.7 + (len(kw) / len(col_name)) * 0.2, "contains"
                elif col_name in kw:
                    score, method = 0.6 + (len(col_name) / len(kw)) * 0.2, "reverse_contains"
                else:
                    score, method = SequenceMatcher(None, col_name, kw).ratio(), "sequence"

                if score > best_score:
                    best_score, best_field, best_method = score, field_name, method

        results.append({
            "col_name": header["name"], "col_index": header["col_index"],
            "suggested_field": best_field, "confidence": round(best_score, 2), "method": best_method,
        })
    return results


# ═══════════════════════════════════════════════════════════
# Step 3: LLM for uncertain columns only (Approach 1)
# ═══════════════════════════════════════════════════════════

def llm_map_uncertain(
    uncertain_columns: list[dict],
    sheet_name: str,
    sample_rows: list[list],
    llm: LLMClient | None = None,
) -> list[dict]:
    """Send ONLY uncertain columns (confidence < 0.8) to LLM.

    Returns list of {col_name, field, confidence, reason}.
    """
    if not uncertain_columns:
        return []

    if not llm or not llm.available:
        # No LLM — return uncertain columns as-is (flagged for human review)
        return [
            {"col_name": c["col_name"], "field": c["suggested_field"] or "unknown",
             "confidence": c["confidence"], "reason": "LLM unavailable, fuzzy only"}
            for c in uncertain_columns
        ]

    # Build prompt with uncertain columns + sample data
    col_descriptions = []
    for c in uncertain_columns:
        col_idx = c["col_index"] - 1
        samples = [row[col_idx] for row in sample_rows if col_idx < len(row) and row[col_idx]][:3]
        col_descriptions.append(
            f'  - "{c["col_name"]}" (fuzzy suggested: {c["suggested_field"]}, '
            f'confidence: {c["confidence"]}) | samples: {samples}'
        )

    prompt = PARSER_COLUMN_MAP_PROMPT.format(
        sheet_name=sheet_name,
        uncertain_columns="\n".join(col_descriptions),
    )

    result = llm.complete_json(prompt, system=PARSER_COLUMN_MAP_SYSTEM, agent_name="Parser Agent")

    if not result or "mappings" not in result:
        # LLM failed — flag everything for human review
        return [
            {"col_name": c["col_name"], "field": "unknown",
             "confidence": 0.0, "reason": "LLM returned no result"}
            for c in uncertain_columns
        ]

    return result["mappings"]


# ═══════════════════════════════════════════════════════════
# Main: ColumnMapper class
# ═══════════════════════════════════════════════════════════

class ColumnMapper:
    """Map columns in any file using Approach 1.

    Flow:
    1. Check DB for saved mappings (instant if same company/format)
    2. Fuzzy match all columns
    3. confidence >= 0.8 → auto-map
    4. confidence < 0.8 → send to LLM
    5. LLM underconfident → flag for human

    Usage:
        mapper = ColumnMapper(repo, llm)
        result = mapper.map_file("file.xlsx", company_id="abc", file_type="tds")
    """

    def __init__(self, repo=None, llm: LLMClient | None = None):
        self.repo = repo
        self.llm = llm

    def map_file(self, filepath: str, company_id: str = "",
                 file_type: str = "auto") -> dict:
        """Map columns in a file. Returns mapping result."""

        # Check saved mappings first
        if company_id and self.repo and file_type != "auto":
            saved = self.repo.column_maps.get_confirmed(company_id, file_type)
            if saved:
                return {"sheets": [{
                    "sheet_name": "saved", "document_type": "saved", "header_row": 0,
                    "mappings": [
                        {"col_name": s.source_column, "field": s.mapped_to,
                         "confidence": 1.0, "source": "saved", "needs_review": False}
                        for s in saved
                    ],
                    "needs_user_review": [], "from_cache": True,
                }]}

        # Read file
        sheets = read_file_headers(filepath)
        result_sheets = []

        for sheet in sheets:
            # Pick field definitions
            fields = TDS_FIELDS if file_type == "tds" else LEDGER_FIELDS
            if file_type == "auto":
                all_headers = " ".join(h["name"].lower() for h in sheet["headers"])
                fields = TDS_FIELDS if any(kw in all_headers for kw in ["section", "tax deducted", "tds"]) else LEDGER_FIELDS

            # Step 1: Fuzzy match all columns
            fuzzy_results = fuzzy_match_columns(sheet["headers"], fields)

            # Step 2: Split by confidence
            auto_mapped = [r for r in fuzzy_results if r["confidence"] >= AUTO_MAP_THRESHOLD]
            uncertain = [r for r in fuzzy_results if r["confidence"] < AUTO_MAP_THRESHOLD]

            # Step 3: Send uncertain to LLM
            llm_results = llm_map_uncertain(uncertain, sheet["sheet_name"], sheet["sample_rows"], self.llm)

            # Build final mappings
            final_mappings = []

            # Auto-mapped (high confidence fuzzy)
            for r in auto_mapped:
                final_mappings.append({
                    "col_name": r["col_name"], "col_index": r["col_index"],
                    "field": r["suggested_field"],
                    "confidence": r["confidence"], "source": "fuzzy_auto",
                    "needs_review": False,
                })

            # LLM-mapped (was uncertain)
            llm_map = {m["col_name"]: m for m in llm_results}
            for r in uncertain:
                llm_m = llm_map.get(r["col_name"], {})
                llm_field = llm_m.get("field", "unknown")
                llm_conf = llm_m.get("confidence", 0)
                llm_reason = llm_m.get("reason", "")

                needs_review = llm_field in ("unknown", None) or llm_conf < 0.6
                final_mappings.append({
                    "col_name": r["col_name"], "col_index": r["col_index"],
                    "field": llm_field if llm_field not in ("unknown", None) else r["suggested_field"],
                    "confidence": llm_conf if llm_conf > 0 else r["confidence"],
                    "source": "llm" if llm_field not in ("unknown", None) else "uncertain",
                    "reason": llm_reason,
                    "needs_review": needs_review,
                })

            needs_review = [m["col_name"] for m in final_mappings if m.get("needs_review")]

            result_sheets.append({
                "sheet_name": sheet["sheet_name"],
                "document_type": "auto_detected",
                "header_row": sheet["header_row"],
                "total_rows": sheet["total_rows"],
                "mappings": final_mappings,
                "needs_user_review": needs_review,
                "stats": {
                    "total_columns": len(fuzzy_results),
                    "auto_mapped": len(auto_mapped),
                    "llm_mapped": len([m for m in final_mappings if m["source"] == "llm"]),
                    "needs_review": len(needs_review),
                },
            })

            # Save confirmed mappings to DB
            if company_id and self.repo:
                ft = "tds" if fields == TDS_FIELDS else "ledger"
                for m in final_mappings:
                    if m["confidence"] >= 0.8 and not m.get("needs_review"):
                        self.repo.column_maps.upsert(
                            company_id=company_id, file_type=ft,
                            source_column=m["col_name"], mapped_to=m["field"],
                            confidence=m["confidence"],
                        )

        return {"sheets": result_sheets}
