"""
Reporter Agent — generates summary, CSV, and Excel reports.

Reads match results and findings from DB, produces:
- reconciliation_summary.json (executive summary)
- reconciliation_report.csv (flat match table)
- findings_report.csv (compliance findings)
- tds_recon_report.xlsx (3-sheet Excel)
"""

import csv
import json
from pathlib import Path

from app.agents.base import AgentBase


class ReporterAgent(AgentBase):
    agent_name = "Reporter Agent"

    def run(self, match_summary: dict, checker_summary: dict,
            matches: list[dict], findings: list[dict],
            output_dir: str = "data/reports") -> dict:
        """Generate all reports.

        Args:
            match_summary: from MatcherAgent.run() return
            checker_summary: from TdsCheckerAgent.run() return
            matches: raw match list from matcher
            findings: raw findings list from checker
            output_dir: where to write report files

        Returns summary dict.
        """
        self.events.agent_start(self.agent_name, "Generating reports...")
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        # Build executive summary
        summary = {
            "run_id": self.run_id,
            "company_id": self.company_id,
            "financial_year": self.financial_year,
            "matching": match_summary,
            "compliance": checker_summary.get("summary", {}),
        }

        # LLM: Generate narrative executive brief
        narrative = None
        if self.llm and self.llm.available:
            self.events.detail(self.agent_name, "Writing executive narrative...")
            narrative = self._generate_narrative(match_summary, checker_summary, findings)
            if narrative:
                summary["narrative"] = narrative
                self.events.emit(self.agent_name, f"Executive brief: {narrative[:150]}...", "llm_insight")

        # Write JSON summary
        summary_file = output_path / "reconciliation_summary.json"
        with open(summary_file, "w") as f:
            json.dump(summary, f, indent=2, default=str)
        self.events.detail(self.agent_name, "reconciliation_summary.json generated")

        # Write CSV match report
        if matches:
            report_file = output_path / "reconciliation_report.csv"
            self._write_match_csv(report_file, matches)
            self.events.detail(self.agent_name, f"reconciliation_report.csv ({len(matches)} rows)")

        # Write CSV findings report
        if findings:
            findings_file = output_path / "findings_report.csv"
            self._write_findings_csv(findings_file, findings)
            self.events.detail(self.agent_name, f"findings_report.csv ({len(findings)} rows)")

        # Write Excel report (with narrative if available)
        try:
            excel_file = output_path / "tds_recon_report.xlsx"
            self._write_excel_report(excel_file, matches, findings, narrative)
            self.events.detail(self.agent_name, "tds_recon_report.xlsx (4 sheets)")
        except Exception as e:
            self.events.warning(self.agent_name, f"Excel generation failed: {e}")

        # Write summary to DB
        self.db.summaries.bulk_insert([{
            "reconciliation_run_id": self.run_id,
            "section": "ALL",
            "group_type": "summary",
            "group_key": "executive_summary",
            "entry_count": match_summary.get("matched", 0),
            "total_amount": match_summary.get("total_resolved", 0),
            "llm_summary": narrative or json.dumps(summary, default=str),
            "status": "resolved" if checker_summary.get("summary", {}).get("errors", 0) == 0 else "needs_attention",
        }])

        # Update run status
        self.db.runs.update_status(self.run_id, "completed")

        self.events.success(self.agent_name, "Reports generated")
        self.events.agent_done(self.agent_name, "Complete")

        return summary

    def _write_match_csv(self, filepath, matches):
        rows = []
        for i, m in enumerate(matches, 1):
            f26 = m.get("form26_entry", {})
            rows.append({
                "sr_no": i,
                "vendor": f26.get("vendor_name", ""),
                "section": f26.get("section", ""),
                "amount": f26.get("amount_paid", 0),
                "tds": f26.get("tax_deducted", 0),
                "match_type": m.get("pass_name", ""),
                "confidence": m.get("confidence", 0),
            })
        if rows:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=rows[0].keys())
                writer.writeheader()
                writer.writerows(rows)

    def _write_findings_csv(self, filepath, findings):
        rows = []
        for i, f in enumerate(findings, 1):
            rows.append({
                "sr_no": i,
                "severity": f.get("severity", ""),
                "check": f.get("check", ""),
                "vendor": f.get("vendor", ""),
                "section": f.get("form26_section", f.get("expected_section", "")),
                "message": f.get("message", ""),
            })
        if rows:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=rows[0].keys())
                writer.writeheader()
                writer.writerows(rows)

    def _write_excel_report(self, filepath, matches, findings, narrative=None):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="467273", end_color="467273", fill_type="solid")

        # Sheet 0: Executive Summary (if narrative available)
        if narrative:
            ws0 = wb.active
            ws0.title = "Executive Summary"
            ws0.cell(row=1, column=1, value="TDS Reconciliation — Executive Summary")
            ws0.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws0.cell(row=2, column=1, value=f"Financial Year: {self.financial_year}")
            ws0.cell(row=3, column=1, value=f"Run ID: {self.run_id}")
            ws0.cell(row=5, column=1, value=narrative)
            ws0.cell(row=5, column=1).alignment = Alignment(wrap_text=True)
            ws0.column_dimensions["A"].width = 100

        # Sheet 1: Issues
        if narrative:
            ws1 = wb.create_sheet("Issues for Review")
        else:
            ws1 = wb.active
            ws1.title = "Issues for Review"
        headers = ["Sr.", "Severity", "Check", "Vendor", "Section", "Finding"]
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for i, f in enumerate(findings, 2):
            ws1.cell(row=i, column=1, value=i - 1)
            ws1.cell(row=i, column=2, value=f.get("severity", ""))
            ws1.cell(row=i, column=3, value=f.get("check", ""))
            ws1.cell(row=i, column=4, value=f.get("vendor", ""))
            ws1.cell(row=i, column=5, value=f.get("form26_section", f.get("expected_section", "")))
            ws1.cell(row=i, column=6, value=f.get("message", ""))

        # Sheet 2: Matched
        ws2 = wb.create_sheet("TDS Report - Matched")
        headers2 = ["Sr.", "Vendor", "Section", "Amount", "TDS", "Match Type", "Confidence"]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for i, m in enumerate(matches, 2):
            f26 = m.get("form26_entry", {})
            ws2.cell(row=i, column=1, value=i - 1)
            ws2.cell(row=i, column=2, value=f26.get("vendor_name", ""))
            ws2.cell(row=i, column=3, value=f26.get("section", ""))
            ws2.cell(row=i, column=4, value=f26.get("amount_paid", 0))
            ws2.cell(row=i, column=5, value=f26.get("tax_deducted", 0))
            ws2.cell(row=i, column=6, value=m.get("pass_name", ""))
            ws2.cell(row=i, column=7, value=m.get("confidence", 0))

        # Sheet 3: Zero TDS
        ws3 = wb.create_sheet("Zero TDS - Exempt")
        headers3 = ["Sr.", "Note"]
        for col, h in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        ws3.cell(row=2, column=1, value=1)
        ws3.cell(row=2, column=2, value="Below-threshold entries are resolved — no TDS required")

        wb.save(str(filepath))
        wb.close()

    def _generate_narrative(self, match_summary: dict, checker_summary: dict,
                            findings: list[dict]) -> str | None:
        """Ask LLM to write a professional narrative summary."""
        from app.services.llm_prompts import REPORTER_NARRATIVE_SYSTEM, REPORTER_NARRATIVE_PROMPT

        # Build matching summary text
        matching_text = (
            f"Total Form 26 entries in scope: {match_summary.get('total_form26', 'N/A')}\n"
            f"Matched with TDS: {match_summary.get('matched', 0)}\n"
            f"Below threshold (TDS=0): {match_summary.get('below_threshold', 0)}\n"
            f"Total resolved: {match_summary.get('total_resolved', 0)}\n"
            f"Unmatched: {match_summary.get('unmatched', 0)}\n"
            f"By pass: {match_summary.get('by_pass', {})}"
        )

        # Build compliance text
        comp = checker_summary.get("summary", {})
        compliance_text = (
            f"Total findings: {comp.get('total', 0)}\n"
            f"Errors: {comp.get('errors', 0)}\n"
            f"Warnings: {comp.get('warnings', 0)}\n"
            f"Exposure: Rs {comp.get('exposure', 0):,.0f}"
        )

        # Top issues
        errors = [f for f in findings if f.get("severity") == "error"]
        top_issues_text = "\n".join(
            f"- {f.get('vendor', 'Unknown')}: {f.get('message', '')[:100]}"
            for f in errors[:5]
        ) or "No critical issues found."

        prompt = REPORTER_NARRATIVE_PROMPT.format(
            financial_year=self.financial_year,
            company_name=self.company_id[:8] + "...",  # will be replaced with real name later
            matching_summary=matching_text,
            compliance_summary=compliance_text,
            top_issues=top_issues_text,
        )

        result = self.llm.complete(prompt, system=REPORTER_NARRATIVE_SYSTEM, agent_name=self.agent_name)
        return result
