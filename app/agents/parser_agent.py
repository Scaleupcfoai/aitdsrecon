"""
Parser Agent — Parse any XLSX/CSV file using the new cascade column mapper.

Uses the L0-L4 cascade (template → exact → fuzzy+fingerprint → LLM)
instead of the old fuzzy-only mapper. Zero hardcoded column positions.

Flow:
1. Excel Loader reads file, finds headers, cleans data
2. Cache check — if same columns seen before, use saved mapping
3. Fingerprinter builds column profiles (dtype, samples, patterns)
4. Cascade matcher identifies columns (L0 template → L1 → L2 → L4)
5. Parser reads rows using confirmed column positions
6. Writes to ledger_entry and tds_entry tables
"""

import re
from datetime import datetime, date

import openpyxl

from app.agents.base import AgentBase
from app.ingestion.excel_loader import load_excel
from app.ingestion.fingerprinter import fingerprint_columns
from app.matching.cache import MappingCache
from app.matching.cascade import CascadeMatcher, MappingResult


# ── Helpers ──

def clean_name(raw_name: str) -> dict:
    """Parse Form 26 name field: 'Adi Debnath (34); PAN: AAAAA0001A' → {name, id, pan}"""
    if not raw_name:
        return {"name": "", "id": "", "pan": ""}
    match = re.match(r"^(.+?)\s*\((\d+)\);\s*PAN:\s*(\S+)", str(raw_name))
    if match:
        return {"name": match.group(1).strip(), "id": match.group(2), "pan": match.group(3)}
    match2 = re.match(r"^(.+?);\s*PAN:\s*(\S+)", str(raw_name))
    if match2:
        return {"name": match2.group(1).strip(), "id": "", "pan": match2.group(2)}
    return {"name": str(raw_name).strip(), "id": "", "pan": ""}


def to_date_str(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, (int, float)):
        serial = int(val)
        if 1 <= serial <= 55000:
            try:
                from datetime import timedelta
                excel_epoch = datetime(1899, 12, 30)
                dt = excel_epoch + timedelta(days=serial)
                return dt.date().isoformat()
            except (ValueError, OverflowError):
                return None
        return None
    if isinstance(val, str):
        val = val.strip()
        if not val or val == "-":
            return None
        if len(val) >= 10 and val[4] == '-':
            return val[:10]
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(val[:10], fmt).date().isoformat()
            except ValueError:
                continue
        return val[:10] if len(val) >= 10 else None
    return None


def safe_float(val) -> float:
    """Convert value to float. Handles None, strings with commas, text, negatives."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        cleaned = val.replace(",", "").replace(" ", "").replace("₹", "").replace("Rs", "").replace("rs", "").strip()
        if not cleaned or cleaned == "-":
            return 0.0
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


# ── Expense type classification ──

def _build_expense_keywords() -> dict:
    try:
        from app.knowledge import get_sections
        keywords = {}
        section_to_type = {
            "194A": "interest_payment", "194C": "freight_expense",
            "194H": "brokerage", "194I_a": "rent", "194I_b": "rent",
            "194J_a": "consultancy", "194J_b": "professional_fees",
            "194D": "insurance", "194Q": "purchase", "192": "salary",
        }
        for code, section in get_sections().items():
            exp_type = section_to_type.get(code, "other")
            if exp_type != "other":
                keywords[exp_type] = section.get("expense_keywords", [])
        keywords["tds_deduction"] = ["tds payable"]
        keywords["salary"] = keywords.get("salary", []) + ["bonus", "director's salary"]
        return keywords
    except Exception:
        return {
            "interest_payment": ["interest paid", "interest on loan"],
            "freight_expense": ["freight", "carriage", "transport", "logistics"],
            "packing_expense": ["packing"],
            "brokerage": ["brokerage", "commission"],
            "rent": ["rent", "shop rent", "office rent"],
            "professional_fees": ["professional", "legal", "audit fees"],
            "salary": ["salary", "bonus", "director's salary"],
            "tds_deduction": ["tds payable"],
            "insurance": ["insurance"],
            "purchase": ["purchase"],
        }


EXPENSE_KEYWORDS = _build_expense_keywords()

EXPENSE_TO_SECTION = {
    "interest_payment": "194A", "freight_expense": "194C", "packing_expense": "194C",
    "brokerage": "194H", "rent": "194I(b)", "consultancy": "194J(b)",
    "professional_fees": "194J(b)", "insurance": "194D", "purchase": "194Q",
}


def _looks_like_number(val: str) -> bool:
    """Check if a string value looks like a number (e.g., '70168.0', '1234')."""
    cleaned = val.strip().replace(",", "").replace(" ", "")
    if not cleaned:
        return False
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def classify_expense(text: str) -> str:
    if not text:
        return "other"
    lower = text.lower()
    for exp_type, keywords in EXPENSE_KEYWORDS.items():
        if any(kw in lower for kw in keywords):
            return exp_type
    return "other"


# ═══════════════════════════════════════════════════════════
# Parser Agent
# ═══════════════════════════════════════════════════════════

class ParserAgent(AgentBase):
    agent_name = "Parser Agent"

    def run(self, form26_path: str, tally_path: str) -> dict:
        """Parse both files using the new cascade column mapper.

        Flow:
        1. Load Excel → find headers, clean data
        2. Check cache → skip cascade if same format seen before
        3. Fingerprint + Cascade → map columns
        4. Build entries from mapped columns
        5. Insert to DB

        Returns: {tds_count, ledger_count, sections, mappings}
        """
        self.events.agent_start(self.agent_name, "Starting Parser Agent...")

        from pathlib import Path
        if not Path(form26_path).exists():
            raise ValueError(f"Form 26 file not found: {form26_path}")
        if not Path(tally_path).exists():
            raise ValueError(f"Tally file not found: {tally_path}")

        # Auto-detect company
        company_info = self._detect_company(form26_path, tally_path)
        if company_info.get("name"):
            self.events.detail(self.agent_name, f"Company detected: {company_info['name']}")
            company = self._ensure_company(company_info)
            if company:
                self.company_id = company.id
                self.events.detail(self.agent_name, f"Company ID: {company.id}")

        cache = MappingCache(self.db)

        # ═══ Parse Form 26 ═══
        self.events.detail(self.agent_name, "Loading Form 26...")
        f26_sheets = load_excel(form26_path)
        f26_mappings = []

        for sheet in f26_sheets:
            mapping = self._map_sheet(sheet, file_type="tds", cache=cache)
            f26_mappings.append({"sheet": sheet, "mapping": mapping})

        # Log Form 26 results
        for m in f26_mappings:
            structural = [r for r in m["mapping"] if r.target and r.target not in ("skip", "gst_column", "expense_head")]
            self.events.detail(self.agent_name,
                f"Form 26 [{m['sheet']['sheet_name']}]: {len(structural)} fields mapped via {m['mapping'][0].method if m['mapping'] else '?'}")

        # ═══ Parse Tally ═══
        self.events.detail(self.agent_name, "Loading Tally extract...")
        tally_sheets = load_excel(tally_path)
        tally_mappings = []

        for sheet in tally_sheets:
            mapping = self._map_sheet(sheet, file_type="ledger", cache=cache)
            tally_mappings.append({"sheet": sheet, "mapping": mapping})

        for m in tally_mappings:
            structural = [r for r in m["mapping"] if r.target and r.target not in ("skip", "gst_column", "expense_head")]
            exp_heads = [r for r in m["mapping"] if r.is_expense_head]
            gst_cols = [r for r in m["mapping"] if r.is_gst_column]
            self.events.detail(self.agent_name,
                f"Tally [{m['sheet']['sheet_name']}]: {len(structural)} fields + {len(exp_heads)} expense heads + {len(gst_cols)} GST cols")

        # ═══ Column Confirmation Gate ═══
        # Pipeline BLOCKS here until user confirms in frontend.
        # If mappings came from DB cache (confirmed=True), skip this step.
        all_from_cache = all(
            m["mapping"][0].method == "cached" if m["mapping"] else True
            for m in f26_mappings + tally_mappings
        )

        if not all_from_cache:
            confirmed = self._confirm_columns_structured(
                f26_mappings, tally_mappings, cache
            )
            if not confirmed:
                self.events.warning(self.agent_name, "Column confirmation timed out — proceeding with detected mappings")

        # ═══ Build entries ═══
        tds_entries = []
        for m in f26_mappings:
            entries = self._build_tds_entries(m["sheet"]["df"], m["mapping"])
            tds_entries.extend(entries)

        # ═══ Self-Validate TDS entries ═══
        tds_issues = self._validate_entries(tds_entries, "tds")
        if tds_issues:
            for issue in tds_issues:
                self.events.warning(self.agent_name, f"TDS data quality: {issue}")
            self.events.emit(self.agent_name,
                f"Parser detected {len(tds_issues)} data quality issue(s) in Form 26 entries. "
                f"This may indicate incorrect column mappings.",
                "warning", data={"validation_issues": tds_issues, "entry_type": "tds"})

        tds_count = self.db.entries.bulk_insert_tds(tds_entries) if tds_entries else 0
        sections = set(e["tds_section"] for e in tds_entries if e.get("tds_section"))
        self.events.detail(self.agent_name, f"Form 26: {tds_count} entries across {len(sections)} sections")

        ledger_entries = []
        for m in tally_mappings:
            entries = self._build_ledger_entries(m["sheet"]["df"], m["mapping"], m["sheet"]["sheet_name"])
            ledger_entries.extend(entries)

        # ═══ Self-Validate Ledger entries ═══
        ledger_issues = self._validate_entries(ledger_entries, "ledger")
        if ledger_issues:
            for issue in ledger_issues:
                self.events.warning(self.agent_name, f"Ledger data quality: {issue}")
            self.events.emit(self.agent_name,
                f"Parser detected {len(ledger_issues)} data quality issue(s) in Tally entries. "
                f"This may indicate incorrect column mappings.",
                "warning", data={"validation_issues": ledger_issues, "entry_type": "ledger"})

        ledger_count = self.db.entries.bulk_insert_ledger(ledger_entries) if ledger_entries else 0
        self.events.detail(self.agent_name, f"Tally: {ledger_count} ledger entries")

        # Combine all issues for the orchestrator to evaluate
        all_issues = tds_issues + ledger_issues

        self.db.runs.update_status(self.run_id, "processing")
        self.events.success(self.agent_name, f"Parsed {tds_count} TDS + {ledger_count} ledger entries")
        if all_issues:
            self.events.warning(self.agent_name,
                f"{len(all_issues)} data quality issue(s) detected — orchestrator will evaluate")
        self.events.agent_done(self.agent_name, "Parsing complete")

        return {
            "tds_count": tds_count,
            "ledger_count": ledger_count,
            "sections": sorted(sections),
            "company": company_info,
            "validation_issues": all_issues,
        }

    # ─── Self-Validation ────────────────────────────────────

    @staticmethod
    def _validate_entries(entries: list[dict], entry_type: str) -> list[str]:
        """Validate parsed entries BEFORE they go to DB.

        Catches column mapping errors that would poison downstream matching.
        Returns list of issue descriptions. Empty = clean data.

        Checks:
        - Critical fields not null/empty beyond threshold
        - Values that look like they came from the wrong column
          (numbers as names, dates as amounts, etc.)
        - Sanity checks on value ranges
        """
        if not entries:
            return [f"0 {entry_type} entries parsed — file may be empty or mapping failed"]

        issues = []
        total = len(entries)

        if entry_type == "tds":
            # ── party_name checks ──
            null_names = sum(1 for e in entries
                            if not e.get("party_name") or str(e["party_name"]).strip() in ("", "None", "nan"))
            if null_names > total * 0.3:
                issues.append(f"party_name missing in {null_names}/{total} entries ({null_names*100//total}%) — likely wrong column mapped")

            # Names that are numbers = reading amount column as name
            numeric_names = sum(1 for e in entries
                               if e.get("party_name") and _looks_like_number(str(e["party_name"])))
            if numeric_names > total * 0.1:
                issues.append(f"party_name contains numbers in {numeric_names}/{total} entries — column swap suspected")

            # ── gross_amount checks ──
            zero_amounts = sum(1 for e in entries if safe_float(e.get("gross_amount")) == 0)
            if zero_amounts > total * 0.3:
                issues.append(f"gross_amount is 0 in {zero_amounts}/{total} entries — likely wrong column or empty data")

            # ── tds_amount checks ──
            zero_tds = sum(1 for e in entries if safe_float(e.get("tds_amount")) == 0)
            if zero_tds > total * 0.5:
                issues.append(f"tds_amount is 0 in {zero_tds}/{total} entries — column mapping may have read date/text as amount")

            # TDS > gross = impossible (TDS is always a fraction of gross)
            tds_gt_gross = sum(1 for e in entries
                              if safe_float(e.get("tds_amount")) > safe_float(e.get("gross_amount")) > 0)
            if tds_gt_gross > total * 0.1:
                issues.append(f"tds_amount > gross_amount in {tds_gt_gross}/{total} entries — columns may be swapped")

            # ── date checks ──
            null_dates = sum(1 for e in entries if not e.get("date_of_deduction"))
            if null_dates > total * 0.5:
                issues.append(f"date_of_deduction missing in {null_dates}/{total} entries — date column may not be mapped")

            # ── section checks ──
            null_sections = sum(1 for e in entries
                               if not e.get("tds_section") or str(e["tds_section"]).strip() in ("", "nan"))
            if null_sections > total * 0.3:
                issues.append(f"tds_section missing in {null_sections}/{total} entries")

        elif entry_type == "ledger":
            # ── party_name checks ──
            null_names = sum(1 for e in entries
                            if not e.get("party_name") or str(e["party_name"]).strip() in ("", "None", "nan"))
            if null_names > total * 0.3:
                issues.append(f"party_name missing in {null_names}/{total} entries ({null_names*100//total}%)")

            numeric_names = sum(1 for e in entries
                               if e.get("party_name") and _looks_like_number(str(e["party_name"])))
            if numeric_names > total * 0.1:
                issues.append(f"party_name contains numbers in {numeric_names}/{total} entries — column swap suspected")

            # ── amount checks ──
            zero_amounts = sum(1 for e in entries if safe_float(e.get("amount")) == 0)
            if zero_amounts > total * 0.5:
                issues.append(f"amount is 0 in {zero_amounts}/{total} entries")

            # ── date checks ──
            null_dates = sum(1 for e in entries if not e.get("invoice_date"))
            if null_dates > total * 0.5:
                issues.append(f"invoice_date missing in {null_dates}/{total} entries")

            # ── GST sanity: GST should not be > amount ──
            gst_gt_amount = sum(1 for e in entries
                               if e.get("gst_amount") and safe_float(e.get("gst_amount")) > safe_float(e.get("amount")) > 0)
            if gst_gt_amount > total * 0.05:
                issues.append(f"gst_amount > amount in {gst_gt_amount}/{total} entries — GST column may be wrong")

        return issues

    # ─── Column confirmation (structured JSON for frontend) ──

    def _confirm_columns_structured(self, f26_mappings, tally_mappings, cache) -> bool:
        """Emit structured column mapping data for frontend confirmation.

        Sends a 'question' event with type='column_confirmation' containing
        structured JSON that the frontend renders as an editable table.

        Pipeline BLOCKS until user confirms via POST /api/answer.
        Returns True if confirmed, False if timed out.
        """
        import uuid

        # Build structured mapping data for frontend
        from app.ingestion.fingerprinter import detect_cross_column_relationships

        confirmation_data = {
            "type": "column_confirmation",
            "company_id": self.company_id,
            "files": [],
            "relationships": [],
        }

        # Form 26 mappings
        for m in f26_mappings:
            df = m["sheet"]["df"]
            file_data = {
                "file_type": "tds",
                "sheet_name": m["sheet"]["sheet_name"],
                "total_rows": m["sheet"]["total_rows"],
                "columns": [],
                "preview_rows": self._get_preview_rows(df, m["mapping"], 3),
            }
            for r in m["mapping"]:
                if r.target == "skip":
                    continue
                file_data["columns"].append({
                    "col_index": r.col_index,
                    "source_name": r.source_name,
                    "mapped_to": r.target,
                    "confidence": r.confidence,
                    "method": r.method,
                    "tier": r.tier,
                    "alternatives": r.alternatives,
                    "sample_values": r.sample_values,
                    "dtype": r.dtype_inferred,
                })

            # Detect cross-column relationships
            from app.ingestion.fingerprinter import fingerprint_columns
            fps = fingerprint_columns(df)
            rels = detect_cross_column_relationships(df, fps)
            rate_rels = [r for r in rels if r["type"] == "rate_product"]
            if rate_rels:
                confirmation_data["relationships"].extend([
                    {"formula": r["formula"], "interpretation": r["interpretation"]}
                    for r in rate_rels
                ])

            confirmation_data["files"].append(file_data)

        # Tally mappings
        for m in tally_mappings:
            df = m["sheet"]["df"]
            file_data = {
                "file_type": "ledger",
                "sheet_name": m["sheet"]["sheet_name"],
                "total_rows": m["sheet"]["total_rows"],
                "columns": [],
                "preview_rows": self._get_preview_rows(df, m["mapping"], 3),
            }
            for r in m["mapping"]:
                if r.target in ("skip",) or r.is_expense_head or r.is_gst_column:
                    continue
                file_data["columns"].append({
                    "col_index": r.col_index,
                    "source_name": r.source_name,
                    "mapped_to": r.target,
                    "confidence": r.confidence,
                    "method": r.method,
                    "tier": r.tier,
                    "alternatives": r.alternatives,
                    "sample_values": r.sample_values,
                    "dtype": r.dtype_inferred,
                })

            gst_count = sum(1 for r in m["mapping"] if r.is_gst_column)
            exp_count = sum(1 for r in m["mapping"] if r.is_expense_head)
            file_data["gst_columns"] = gst_count
            file_data["expense_head_columns"] = exp_count

            confirmation_data["files"].append(file_data)

        # Build human-readable summary for the question message
        total_cols = sum(len(f["columns"]) for f in confirmation_data["files"])
        high_conf = sum(
            1 for f in confirmation_data["files"]
            for c in f["columns"] if c["tier"] == "HIGH"
        )
        needs_review = total_cols - high_conf

        message = f"I've detected column mappings for {len(confirmation_data['files'])} sheets ({total_cols} columns). "
        if needs_review > 0:
            message += f"{needs_review} column(s) need your review. "
        else:
            message += "All columns mapped with high confidence. "
        message += "Please confirm before I proceed with parsing."

        q_id = f"q_cols_{uuid.uuid4().hex[:8]}"

        answer = self.events.question(
            agent=self.agent_name,
            message=message,
            question_id=q_id,
            options=[
                {"id": "confirm", "label": "Confirm & Parse",
                 "description": "These column mappings look correct"},
                {"id": "proceed", "label": "Proceed without review",
                 "description": "Skip review, use detected mappings"},
            ],
            allow_text_input=False,
            multi_select=False,
        )

        # Attach the structured data to the question event
        # The frontend reads this from event.data to render the table
        self.events.emit(self.agent_name, "Column mappings ready for review", "detail",
                        data=confirmation_data)

        if answer:
            selected = answer.get("selected", [])
            confirmed_mappings = answer.get("confirmed_mappings")

            if confirmed_mappings:
                # User corrected some mappings in the frontend — save to DB
                for file_mapping in confirmed_mappings:
                    cache.save(
                        company_id=file_mapping.get("company_id", self.company_id),
                        file_type=file_mapping.get("file_type", "ledger"),
                        mappings=file_mapping.get("columns", []),
                    )
                self.events.detail(self.agent_name, "User confirmed columns — mappings saved to DB")
            elif "confirm" in selected:
                # User accepted auto-detected mappings — save to DB as-is
                for m in f26_mappings:
                    cache.save(self.company_id, "tds",
                              [r.to_dict() for r in m["mapping"] if r.target not in ("skip", None)])
                for m in tally_mappings:
                    cache.save(self.company_id, "ledger",
                              [r.to_dict() for r in m["mapping"]
                               if r.target not in ("skip", None) and not r.is_expense_head and not r.is_gst_column])
                self.events.detail(self.agent_name, "User confirmed columns — mappings saved to DB")
            else:
                self.events.detail(self.agent_name, "User chose to proceed without saving")

            return True

        return False  # Timed out

    @staticmethod
    def _get_preview_rows(df, mapping: list, n: int = 3) -> list[dict]:
        """Get N preview rows using the proposed mapping.

        Returns data as the USER would see it — mapped field names as keys,
        actual cell values. This lets the user visually verify the mapping
        before confirming.
        """
        # Build field→column lookup (only mapped, non-skip fields)
        field_to_col = {}
        for m in mapping:
            target = m.target if hasattr(m, 'target') else m.get('target')
            source = m.source_name if hasattr(m, 'source_name') else m.get('source_name')
            if target and target not in ("skip", "gst_column", "expense_head"):
                is_exp = getattr(m, 'is_expense_head', False) if hasattr(m, 'is_expense_head') else m.get('is_expense_head', False)
                is_gst = getattr(m, 'is_gst_column', False) if hasattr(m, 'is_gst_column') else m.get('is_gst_column', False)
                if not is_exp and not is_gst:
                    field_to_col[target] = source

        if not field_to_col:
            return []

        rows = []
        for _, row in df.head(n).iterrows():
            preview = {}
            for field, col_name in field_to_col.items():
                val = row.get(col_name)
                if val is None or (isinstance(val, float) and str(val) == 'nan'):
                    preview[field] = ""
                elif hasattr(val, 'isoformat'):
                    preview[field] = str(val)[:10]
                else:
                    preview[field] = str(val)[:50]
            rows.append(preview)

        return rows

    # ─── Column mapping per sheet ────────────────────────────

    def _map_sheet(self, sheet: dict, file_type: str, cache: MappingCache) -> list[MappingResult]:
        """Map columns for a single sheet: DB cache → fingerprint → cascade."""
        df = sheet["df"]

        # Check DB cache first (confirmed mappings from previous runs)
        cached = cache.lookup(self.company_id, file_type)
        if cached:
            self.events.detail(self.agent_name,
                f"Cache hit for [{sheet['sheet_name']}] — using confirmed mappings from DB")
            # Convert DB cache format to MappingResult
            # Match cached source_column names to current DataFrame columns
            col_names = list(df.columns)
            results = []
            cache_map = {c["source"]: c for c in cached}
            for idx, col_name in enumerate(col_names):
                c = cache_map.get(col_name)
                if c:
                    results.append(MappingResult(
                        source_name=col_name, col_index=idx,
                        target=c["target"], confidence=c.get("confidence", 1.0),
                        method="cached", tier="HIGH",
                        reason="Confirmed mapping from DB",
                    ))
                else:
                    # Column exists in file but not in cache — new column
                    results.append(MappingResult(
                        source_name=col_name, col_index=idx,
                        target="expense_head" if file_type == "ledger" else None,
                        confidence=0.5, method="unmatched", tier="LOW",
                        reason="Not in saved mappings — new column",
                        is_expense_head=(file_type == "ledger"),
                    ))
            return results

        # No cache → fingerprint + cascade
        fingerprints = fingerprint_columns(df)
        matcher = CascadeMatcher(
            fingerprints=fingerprints,
            file_type=file_type,
            llm=self.llm,
            events=self.events,
        )
        return matcher.match()

    # ─── Build TDS entries from DataFrame ────────────────────

    def _build_tds_entries(self, df, mapping: list[MappingResult]) -> list[dict]:
        """Build tds_entry dicts from DataFrame using cascade mapping results."""
        # Build field → column_name lookup
        field_to_col = {}
        for m in mapping:
            if m.target and m.target not in ("skip", "gst_column", "expense_head"):
                field_to_col[m.target] = m.source_name

        if "party_name" not in field_to_col:
            return []

        entries = []
        for _, row in df.iterrows():
            party_raw = row.get(field_to_col.get("party_name", ""), "")
            if not party_raw or str(party_raw).strip() == "":
                continue

            parsed = clean_name(str(party_raw))
            section = str(row.get(field_to_col.get("tds_section", ""), "")).strip()
            if not section:
                continue  # Skip rows without section (continuation rows, subtotals)

            entry = {
                "reconciliation_run_id": self.run_id,
                "company_id": self.company_id,
                "financial_year": self.financial_year,
                "party_name": parsed["name"],
                "pan": parsed["pan"],
                "tds_section": section,
                "tds_amount": safe_float(row.get(field_to_col.get("tds_amount", ""))),
                "gross_amount": safe_float(row.get(field_to_col.get("gross_amount", ""))),
                "date_of_deduction": to_date_str(row.get(field_to_col.get("date_of_deduction", ""))),
                "raw_data": {
                    "vendor_id": parsed["id"],
                    "tax_rate_pct": safe_float(row.get(field_to_col.get("tax_rate", ""))),
                },
            }
            entries.append(entry)

        return entries

    # ─── Build Ledger entries from DataFrame ─────────────────

    def _build_ledger_entries(self, df, mapping: list[MappingResult], sheet_name: str) -> list[dict]:
        """Build ledger_entry dicts from DataFrame using cascade mapping results."""
        # Build lookups
        field_to_col = {}
        gst_columns = []
        expense_head_columns = []

        for m in mapping:
            if m.is_gst_column:
                gst_columns.append(m.source_name)
            elif m.is_expense_head:
                expense_head_columns.append(m.source_name)
            elif m.target and m.target not in ("skip",):
                field_to_col[m.target] = m.source_name

        if "party_name" not in field_to_col:
            return []

        entries = []
        for _, row in df.iterrows():
            party = str(row.get(field_to_col.get("party_name", ""), "")).strip()
            if not party or party in ("None", "nan", ""):
                continue

            # Skip total rows
            if any(kw in party.lower() for kw in ["total", "grand total"]):
                continue

            # Get amount from mapped field
            amount_col = field_to_col.get("amount") or field_to_col.get("gross_total")
            amount = safe_float(row.get(amount_col, 0)) if amount_col else 0.0

            # Collect GST amounts
            total_gst = 0.0
            gst_breakup = {}
            for gc in gst_columns:
                val = safe_float(row.get(gc, 0))
                if val > 0:
                    gst_breakup[gc] = val
                    total_gst += val

            # Collect expense heads (non-zero values)
            expense_heads = {}
            for ec in expense_head_columns:
                val = safe_float(row.get(ec, 0))
                if val != 0:
                    expense_heads[ec] = val

            # If no mapped amount, try summing expense heads
            if amount == 0 and expense_heads:
                amount = sum(expense_heads.values())

            # Classify expense type from non-zero expense heads
            exp_type = "other"
            for head_name in expense_heads:
                classified = classify_expense(head_name)
                if classified != "other":
                    exp_type = classified
                    break

            tds_section = EXPENSE_TO_SECTION.get(exp_type)

            # Base amount = gross minus GST
            base_amount = amount - total_gst if total_gst > 0 else amount

            entry = {
                "reconciliation_run_id": self.run_id,
                "company_id": self.company_id,
                "financial_year": self.financial_year,
                "party_name": party,
                "expense_type": exp_type,
                "amount": amount,
                "gst_amount": total_gst if total_gst > 0 else None,
                "tds_section": tds_section,
                "invoice_number": str(row.get(field_to_col.get("invoice_number", ""), "")).strip() or None,
                "invoice_date": to_date_str(row.get(field_to_col.get("invoice_date", ""))),
                "raw_data": {
                    "source_sheet": sheet_name,
                    "expense_heads": expense_heads if expense_heads else None,
                    "gst_breakup": gst_breakup if gst_breakup else None,
                    "base_amount": base_amount if total_gst > 0 else None,
                },
            }
            entries.append(entry)

        return entries

    # ─── Company detection (unchanged) ───────────────────────

    def _detect_company(self, form26_path: str, tally_path: str) -> dict:
        """Read company name from file headers."""
        from pathlib import Path
        info = {"name": "", "form_type": "", "period": "", "cin": ""}

        try:
            if Path(form26_path).suffix.lower() in (".xlsx", ".xls"):
                wb = openpyxl.load_workbook(form26_path, data_only=True)
                ws = wb[wb.sheetnames[0]]
                for col in range(1, 10):
                    val = ws.cell(1, col).value
                    if val and isinstance(val, str) and len(val.strip()) > 2:
                        info["name"] = val.strip()
                        break
                for col in range(1, 10):
                    val = ws.cell(2, col).value
                    if val and isinstance(val, str) and "form" in val.lower():
                        info["period"] = val.strip()
                        if "26Q" in val:
                            info["form_type"] = "26Q"
                        break
                wb.close()
        except Exception:
            pass

        if not info["name"]:
            try:
                if Path(tally_path).suffix.lower() in (".xlsx", ".xls"):
                    wb = openpyxl.load_workbook(tally_path, data_only=True)
                    ws = wb[wb.sheetnames[0]]
                    val = ws.cell(1, 1).value
                    if val and isinstance(val, str):
                        info["name"] = val.strip()
                    wb.close()
            except Exception:
                pass

        return info

    def _ensure_company(self, company_info: dict):
        """Create company in DB if it doesn't already exist."""
        name = company_info.get("name", "")
        if not name:
            return None
        try:
            existing = self.db.companies.list_by_firm(self.firm_id)
            for c in existing:
                if c.company_name.lower().strip() == name.lower().strip():
                    return c
            company = self.db.companies.create(
                firm_id=self.firm_id, company_name=name,
                pan="", company_type="company",
            )
            self.events.detail(self.agent_name, f"Auto-created company: {name}")
            return company
        except Exception as e:
            self.events.warning(self.agent_name, f"Could not auto-create company: {str(e)[:50]}")
            return None
