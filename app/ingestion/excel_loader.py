"""
Excel Loader — Load XLSX/CSV, find header row, clean data.

CA firm files often have 3-5 rows of metadata above the actual data.
This loader handles that by scanning for the header row.

Usage:
    from app.ingestion.excel_loader import load_excel
    sheets = load_excel("path/to/file.xlsx")
    # Returns: [{df, header_row, sheet_name, total_rows, metadata}, ...]
"""

import re
from pathlib import Path

import openpyxl
import pandas as pd


# Known financial terms — used to identify header rows
HEADER_SEED_TERMS = [
    "date", "amount", "name", "invoice", "voucher", "tds", "pan",
    "rate", "total", "section", "particulars", "value", "tax",
    "deducted", "deduction", "party", "vendor", "gross",
]


def load_excel(filepath: str) -> list[dict]:
    """Load an XLSX or CSV file. Returns list of sheet results.

    Each result:
    {
        "df": pandas DataFrame (cleaned),
        "header_row": int (0-indexed row in original file),
        "sheet_name": str,
        "total_rows": int (data rows, excluding header),
        "metadata": dict (company name, period, etc. from rows above header),
        "raw_headers": list[str] (original column names before cleaning),
    }
    """
    path = Path(filepath)
    if path.suffix.lower() in (".xlsx", ".xls"):
        return _load_xlsx(filepath)
    elif path.suffix.lower() == ".csv":
        return [_load_csv(filepath)]
    else:
        raise ValueError(f"Unsupported file type: {path.suffix}")


def _load_xlsx(filepath: str) -> list[dict]:
    """Load all sheets from an XLSX file."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip empty or tiny sheets
        if ws.max_row < 3 or ws.max_column < 3:
            continue

        # Step 1: Find header row
        header_row, header_values = _find_header_row(ws)
        if header_row is None:
            continue

        # Step 2: Extract metadata from rows above header
        metadata = _extract_metadata(ws, header_row)

        # Step 3: Clean column names
        raw_headers = []
        clean_headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(header_row, col).value
            raw = str(val).strip() if val else ""
            raw_headers.append(raw)

            # Clean: replace \n with space, strip whitespace
            cleaned = raw.replace("\n", " ").replace("\r", " ").strip()
            # Collapse multiple spaces
            cleaned = re.sub(r"\s+", " ", cleaned)
            clean_headers.append(cleaned)

        # Step 4: Read data rows into list of dicts
        data_rows = []
        for row_num in range(header_row + 1, ws.max_row + 1):
            row_vals = {}
            is_total_row = False
            for col_idx, col_name in enumerate(clean_headers):
                if not col_name:
                    continue
                val = ws.cell(row_num, col_idx + 1).value

                # Check for total/grand total rows
                if col_idx < 2 and val and isinstance(val, str):
                    lower_val = val.strip().lower()
                    if lower_val in ("total", "grand total", "sub total", "sub-total"):
                        is_total_row = True
                        break

                row_vals[col_name] = val

            if is_total_row:
                continue
            # Skip fully empty rows
            if all(v is None or v == "" for v in row_vals.values()):
                continue
            data_rows.append(row_vals)

        if not data_rows:
            continue

        # Step 5: Build DataFrame
        df = pd.DataFrame(data_rows)

        # Drop fully empty columns
        df = df.dropna(axis=1, how="all")

        # Drop columns with empty string names
        df = df[[c for c in df.columns if c.strip()]]

        results.append({
            "df": df,
            "header_row": header_row,
            "sheet_name": sheet_name,
            "total_rows": len(df),
            "metadata": metadata,
            "raw_headers": raw_headers,
        })

    wb.close()
    return results


def _find_header_row(ws) -> tuple[int | None, list[str]]:
    """Find the header row by scanning first 15 rows.

    Header row = row with most text cells where at least one cell
    fuzzy-matches a known financial term.
    """
    best_row = None
    best_score = 0
    best_headers = []

    for row_num in range(1, min(16, ws.max_row + 1)):
        text_cells = []
        term_matches = 0

        for col in range(1, min(ws.max_column + 1, 100)):
            val = ws.cell(row_num, col).value
            if val is None:
                continue
            text = str(val).strip().lower().replace("\n", " ")

            # Skip if it's purely numeric
            if text.replace(".", "").replace(",", "").replace("-", "").isdigit():
                continue

            text_cells.append(text)

            # Check if any seed term is in this cell
            for term in HEADER_SEED_TERMS:
                if term in text:
                    term_matches += 1
                    break

        # Score: number of text cells + bonus for term matches
        # Must have at least 3 text cells and 1 term match
        if len(text_cells) >= 3 and term_matches >= 1:
            score = len(text_cells) + term_matches * 2
            if score > best_score:
                best_score = score
                best_row = row_num
                best_headers = text_cells

    return best_row, best_headers


def _extract_metadata(ws, header_row: int) -> dict:
    """Extract metadata from rows above the header (company name, period, etc.)."""
    metadata = {}

    for row_num in range(1, header_row):
        for col in range(1, min(ws.max_column + 1, 10)):
            val = ws.cell(row_num, col).value
            if val and isinstance(val, str) and len(val.strip()) > 2:
                text = val.strip()

                # Company name is usually the first substantial text
                if "company_name" not in metadata and not any(
                    kw in text.lower() for kw in ["cin:", "e-mail", "print date", "form-"]
                ):
                    metadata["company_name"] = text

                # Detect form type
                if "form" in text.lower() and ("26" in text or "24" in text):
                    metadata["form_type"] = text

                # Detect CIN
                if "cin" in text.lower():
                    metadata["cin"] = text.replace("CIN:", "").replace("CIN", "").strip()

                # Detect period
                if "to" in text.lower() and any(c.isdigit() for c in text):
                    metadata["period"] = text

                # Detect register name
                if "register" in text.lower():
                    metadata["register_name"] = text

                break  # Only first non-empty cell per row

    return metadata


def _load_csv(filepath: str) -> dict:
    """Load a CSV file with header detection."""
    # Read first 10 rows to find header
    with open(filepath, encoding="utf-8-sig") as f:
        lines = []
        for i, line in enumerate(f):
            if i >= 10:
                break
            lines.append(line)

    # Find header row
    best_row = 0
    best_score = 0
    for i, line in enumerate(lines):
        cells = line.strip().split(",")
        text_cells = [c for c in cells if c.strip() and not c.strip().replace(".", "").isdigit()]
        term_matches = sum(1 for c in text_cells if any(t in c.lower() for t in HEADER_SEED_TERMS))
        score = len(text_cells) + term_matches * 2
        if score > best_score and len(text_cells) >= 3:
            best_score = score
            best_row = i

    df = pd.read_csv(filepath, header=best_row, encoding="utf-8-sig")

    # Clean column names
    df.columns = [
        re.sub(r"\s+", " ", str(c).replace("\n", " ").strip())
        for c in df.columns
    ]

    # Drop total rows
    first_col = df.columns[0]
    if df[first_col].dtype == object:
        df = df[~df[first_col].str.lower().str.strip().isin(
            ["total", "grand total", "sub total", "sub-total"]
        ).fillna(False)]

    df = df.dropna(axis=1, how="all")
    df = df.dropna(how="all")

    return {
        "df": df,
        "header_row": best_row,
        "sheet_name": "CSV",
        "total_rows": len(df),
        "metadata": {},
        "raw_headers": list(df.columns),
    }
