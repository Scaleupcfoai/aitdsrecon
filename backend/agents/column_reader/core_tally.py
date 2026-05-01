"""Tally-register extraction, ported from aitdsrecon/tds-recon/agents/parser_agent.py.

A Tally export workbook has three sheet *shapes*:

  1. journal              Row per journal entry. Amount spread across 80+ expense-head
                          columns (Freight, Brokerage, Audit Fees, Director's Salary,
                          named directors, X (Loan) interest, GST ledgers, etc.).
  2. purchase_gst_exp     Row per expense voucher with GST breakup. Base amount =
                          sum of expense-head columns (NOT "Value" or "Gross Total",
                          which are GST-inclusive totals).
  3. purchase_plain       Row per goods-purchase voucher. Value = taxable base;
                          Gross Total = Value + GST.

Sheet names are NOT hardcoded. We detect type from the row-6 header signature.

Every sheet follows this anatomy:
  rows 1..5  = company/title metadata (skip)
  row 6      = column headers
  row 7..N   = data

For each data row, we expand into one or more synthetic "expense rows"
  (one per expense-head column that has a non-zero value), each shaped as:

  {
    "date":        original voucher date
    "vendor":      Particulars (vendor short code from the sheet)
    "pan":         ""            (Tally exports don't carry PAN)
    "amount":      the non-zero cell value for this expense-head column
    "description": the expense-head column name (used for classification)
    "source":      "journal" | "purchase_gst_exp" | "purchase_plain"
    "voucher_no":  voucher number
    "section_hint": section picked from column-name keyword (or None)
  }

These rows drop into b2's calculate_batch unchanged.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

# ── Columns to EXCLUDE from row expansion ──
# These are meta, GST ledgers, TDS liability, non-TDS expense types (Depreciation,
# TCS collected, Income Tax, Salary to employees/directors which fall under 192 slab,
# owner equity etc).
META_COLS = {
    "Date", "Particulars", "Voucher No.", "Value", "Gross Total",
    "Addl. Cost", "Rounded (+/-)",
}

GST_LEDGER_COLS = {
    "C GST", "S GST", "I GST", "Input C GST", "Input S GST", "Input I GST",
    "Payable C GST", "Payable S GST", "GST Paid",
}

# Liabilities / non-TDS buckets / out-of-scope categories.
EXCLUDE_COLS = {
    "TDS Payable", "TCS on Purchase",
    "Provision of Income Tax", "Advance Tax (22-23)", "Income Tax Refund",
    "TDS Receivable",
    "Depreciation",
    "Purchase Account", "Discount Received From Purchase",
    "Cash Discount", "Discount Without GST (CD)", "Rate Difference Purchase",
    "Rate Difference Purchase Without GST",
    "Sales Pramotions", "Sales Pramotions Without GST",
    "Purchase Pramotions (Gift)",
    "Employees Professonal Tax",
    "Accrued Interest", "Interest Received", "Interest Received From KMB",
    "Income Tax Refund",
    "Insurance for Goods",
    "Misc. Account", "Other Charges",
    "Outstanding Audit Fees",  # accrual only; Audit Fees covers the expense
}

# Substring patterns that mark a column as non-expense regardless of exact name.
# Catches Tally variants like "TDS On Purchase 194Q", "TCS On Sales", "Output CGST",
# "Federal Bank FD:14", "Exchange Gain/Loss F/Y 2024-25", "GST Refundable", etc.
EXCLUDE_PATTERNS = (
    "tds on ", "tds payable", "tds receivable",
    "tcs on", "tcs payable",
    "output cgst", "output sgst", "output igst",
    "input cgst", "input sgst", "input igst",
    "gst payable", "gst paid", "gst refundable", "gst unclaim",
    "exchange gain", "exchange loss", "forex gain", "forex loss",
    "fd:", "mfd",  # Federal Bank FD:N, MFD:N (investment ledgers)
    "fixed deposit", "mutual fund",
    "corpus fund",
    "advance tax", "provision of income tax", "income tax refund",
    "tds receivable",
    "outstanding ",  # outstanding audit fees, etc — accruals
    "running account", "running a/c",
    "ddb incentive", "duty drawback", "rodtep", "meis", "seis",
    "foreign inward remittance",
    "rebate received", "incentive received",
)

# Salary (section 192 — slab-based, different framework) is out of scope for
# this calculator. We flag separately but don't emit TDS rows.
SALARY_COLS = {
    "Salary & Bonus", "Director's Salary", "Incentive Paid",
}

# Any column ending with "(Loan)" is treated as a loan-interest target party —
# its presence alongside "Interest Paid" signals 194A.
LOAN_SUFFIX = "(Loan)"

# Asset / fixed-asset ledgers (not TDS-applicable expenses; they're
# capitalised, not claimed).
FIXED_ASSET_COLS = {
    "Motor Car", "Air Conditioner", "Camera Set", "Furniture & Fixtures",
    "Mobile Set", "Computer",  # when it appears in Journal Register as a ledger
    "Trademark Registration Fees",
}

# ── Column-name → TDS section hint ──
# Used to pre-classify before b2 runs its own classifier.
# None means "let b2 figure it out from description".
SECTION_HINTS = {
    # 194A
    "interest paid": "194A",
    # 194C
    "freight": "194C", "carriage": "194C", "packing": "194C",
    "printing": "194C", "stationary": "194C", "shop repair": "194C",
    "maintenance": "194C", "contractor": "194C",
    # 194H
    "brokerage": "194H", "commission": "194H",
    # 194J(b)
    "professonal charges": "194J(b)", "professional charges": "194J(b)",
    "consultancy": "194J(b)", "audit fees": "194J(b)", "legal": "194J(b)",
    "software": "194J(b)", "domain": "194J(b)",
    "gst return filling": "194J(b)", "gst annual return": "194J(b)",
    # 194I (rent) — we don't have 194I in rates.py yet; leave unmapped
    # so b2's classifier emits an "unknown_expense" flag for user confirmation.
    # Advertisement is intentionally ambiguous — b2 will flag.
}


def _section_hint_for(col_name: str) -> str | None:
    name = col_name.lower()
    for keyword, section in SECTION_HINTS.items():
        if keyword in name:
            return section
    return None


# ── Sheet-type detection ─────────────────────────────────────────────────

def sniff_sheet_type(header_row_values: list[Any]) -> str:
    """Classify a sheet by its row-6 headers.

    Returns one of: 'journal', 'purchase_gst_exp', 'purchase_plain', 'flat', 'unknown'.
    """
    headers = {str(v).strip() for v in header_row_values if v}
    has_particulars = "Particulars" in headers
    has_voucher = "Voucher No." in headers
    has_value = "Value" in headers
    has_gross = "Gross Total" in headers
    has_purchase_acct = "Purchase Account" in headers
    has_addl_cost = "Addl. Cost" in headers

    if not (has_particulars and has_voucher):
        return "flat"

    # Purchase Register has a "Purchase Account" column and ~18 columns.
    if has_purchase_acct and has_value:
        return "purchase_plain"

    # Purchase GST Exp has "Addl. Cost" + moderate column count.
    if has_addl_cost and has_value and has_gross:
        return "purchase_gst_exp"

    # Journal has Value + Gross Total but no Purchase Account / Addl. Cost.
    if has_gross and has_particulars and has_voucher:
        return "journal"

    return "unknown"


# ── Header/data row locators ─────────────────────────────────────────────

def find_header_row(ws, scan_rows: int = 15) -> int:
    """Locate the real header row (1-based).

    Heuristic: the first row whose first cell is 'Date' and cell B is 'Particulars'.
    Falls back to row 6 (Tally default) if not found.
    """
    for r in range(1, scan_rows + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a and str(a).strip().lower() == "date" and b and "particular" in str(b).lower():
            return r
    return 6


def _header_map(ws, header_row: int) -> tuple[dict[str, int], dict[int, str]]:
    """Return (header_name -> col_idx, col_idx -> header_name) for non-empty headers."""
    name_to_idx: dict[str, int] = {}
    idx_to_name: dict[int, str] = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        name = str(cell.value).strip()
        if not name:
            continue
        name_to_idx[name] = cell.column  # 1-based
        idx_to_name[cell.column] = name
    return name_to_idx, idx_to_name


def _iso_date(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, (datetime, date)):
        return (v.date() if isinstance(v, datetime) else v).isoformat()
    try:
        return datetime.fromisoformat(str(v)).date().isoformat()
    except ValueError:
        return str(v)


# ── Row expansion per sheet type ─────────────────────────────────────────

def _skip_col(col_name: str) -> bool:
    if col_name in META_COLS or col_name in GST_LEDGER_COLS:
        return True
    if col_name in EXCLUDE_COLS or col_name in FIXED_ASSET_COLS:
        return True
    # Substring patterns — catch variants like "TDS On Purchase 194Q",
    # "Federal Bank FD:14", "Exchange Gain/Loss F/Y 2024-25", "DDB Incentive Scheme".
    low = col_name.lower()
    for pat in EXCLUDE_PATTERNS:
        if pat in low:
            return True
    # Salary handled separately (we emit a non-TDS 'skip' row, not an expense row).
    if col_name in SALARY_COLS:
        return True
    # Individual director names as columns (from Journal Register): if it's a
    # single-name column (not "(Loan)") and in the known director list, skip.
    # We don't have that list here, so rely on a weaker heuristic: short
    # (<=6 chars) all-caps tokens are director short-codes in the HPC file
    # (AKB, VA, PKS, SKM, SSCB, STK, HT, SKY, AD, ADD, etc.). Treat as
    # director salary / drawing — skip.
    stripped = col_name.strip()
    if stripped.endswith(LOAN_SUFFIX):
        return False  # loan-party column is meaningful (interest under 194A)
    if len(stripped) <= 6 and stripped.isupper():
        return True
    # "X (D1)" / "X (D2)" / "ADD (D3)" — directors' drawings. Skip.
    if "(D" in stripped and stripped.endswith(")"):
        return True
    return False


def _extract_journal(ws, header_row: int) -> list[dict[str, Any]]:
    _, idx_to_name = _header_map(ws, header_row)
    rows: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        date_val = r[0].value
        particulars = r[1].value
        voucher_no = r[2].value
        if (particulars and "grand total" in str(particulars).lower()):
            continue
        if date_val is None and particulars is None:
            continue

        # Collect non-zero, in-scope expense-head columns.
        for cell in r:
            col_name = idx_to_name.get(cell.column)
            if not col_name or _skip_col(col_name):
                continue
            val = cell.value
            if val is None or val == 0:
                continue
            try:
                amount = float(val)
            except (TypeError, ValueError):
                continue
            if amount <= 0:
                continue
            rows.append({
                "date": _iso_date(date_val),
                "vendor": str(particulars or "").strip(),
                "pan": "",
                "amount": round(amount, 2),
                "description": col_name,
                "source": "journal",
                "voucher_no": str(voucher_no or "").strip(),
                "section_hint": _section_hint_for(col_name),
            })
    return rows


def _extract_purchase_gst_exp(ws, header_row: int) -> list[dict[str, Any]]:
    _, idx_to_name = _header_map(ws, header_row)
    rows: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        date_val = r[0].value
        particulars = r[1].value
        voucher_no = r[2].value
        if (particulars and "grand total" in str(particulars).lower()):
            continue
        if date_val is None and particulars is None:
            continue

        for cell in r:
            col_name = idx_to_name.get(cell.column)
            if not col_name or _skip_col(col_name):
                continue
            val = cell.value
            if val is None or val == 0:
                continue
            try:
                amount = float(val)
            except (TypeError, ValueError):
                continue
            if amount <= 0:
                continue
            rows.append({
                "date": _iso_date(date_val),
                "vendor": str(particulars or "").strip(),
                "pan": "",
                "amount": round(amount, 2),
                "description": col_name,
                "source": "purchase_gst_exp",
                "voucher_no": str(voucher_no or "").strip(),
                "section_hint": _section_hint_for(col_name),
            })
    return rows


def _extract_purchase_plain(ws, header_row: int) -> list[dict[str, Any]]:
    """Goods purchases. One row per voucher. Amount = Value (taxable base)."""
    name_to_idx, idx_to_name = _header_map(ws, header_row)
    value_col = name_to_idx.get("Value")
    rows: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        date_val = r[0].value
        particulars = r[1].value
        voucher_no = r[2].value
        if (particulars and "grand total" in str(particulars).lower()):
            continue
        if date_val is None and particulars is None:
            continue

        value = r[value_col - 1].value if value_col else None
        if value is None or value == 0:
            continue
        try:
            amount = float(value)
        except (TypeError, ValueError):
            continue
        if amount <= 0:
            continue

        rows.append({
            "date": _iso_date(date_val),
            "vendor": str(particulars or "").strip(),
            "pan": "",
            "amount": round(amount, 2),
            "description": "Goods Purchase",
            "source": "purchase_plain",
            "voucher_no": str(voucher_no or "").strip(),
            "section_hint": "194Q",
        })
    return rows


# ── Public API ───────────────────────────────────────────────────────────

def list_sheets(path: str) -> list[dict[str, Any]]:
    """Sheet names + dimensions."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for name in wb.sheetnames:
        ws = wb[name]
        out.append({
            "name": name,
            "rows": ws.max_row,
            "cols": ws.max_column,
        })
    wb.close()
    return out


def sniff_sheet(path: str, sheet_name: str) -> dict[str, Any]:
    """Classify a sheet + return its header row index, headers, and 3 sample rows."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    header_row = find_header_row(ws)
    headers = [
        str(c.value).strip() if c.value else ""
        for c in ws[header_row]
    ]
    samples: list[list[Any]] = []
    for r in ws.iter_rows(min_row=header_row + 1, max_row=header_row + 3, values_only=True):
        samples.append([
            str(v) if v is not None and v != "" else ""
            for v in r
        ])
    sheet_type = sniff_sheet_type(headers)
    wb.close()
    return {
        "sheet": sheet_name,
        "type": sheet_type,
        "header_row": header_row,
        "headers": [h for h in headers if h],
        "samples": samples,
    }


def extract_tally_rows(path: str, sheet_name: str, sheet_type: str) -> dict[str, Any]:
    """Extract normalized expense rows from a single sheet.

    Returns:
      {
        "sheet": ...,
        "sheet_type": ...,
        "row_count": N,
        "rows": [ ... ],          # normalized rows
        "skipped_cols": [ ... ],  # columns skipped with reason (for debugging)
      }
    """
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = wb[sheet_name]
    header_row = find_header_row(ws)

    if sheet_type == "journal":
        rows = _extract_journal(ws, header_row)
    elif sheet_type == "purchase_gst_exp":
        rows = _extract_purchase_gst_exp(ws, header_row)
    elif sheet_type == "purchase_plain":
        rows = _extract_purchase_plain(ws, header_row)
    else:
        rows = []

    wb.close()
    return {
        "sheet": sheet_name,
        "sheet_type": sheet_type,
        "row_count": len(rows),
        "rows": rows,
    }
